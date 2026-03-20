from __future__ import annotations

import base64
import hashlib
import io
import json
import re
import shutil
import unicodedata
import uuid
import zipfile
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
import pytesseract
from openpyxl import load_workbook
from pdf2image import convert_from_path
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
from pypdf import PdfReader, PdfWriter

from .config import settings
from .legacy_bridge import generate_legacy_flatfile_926, generate_legacy_flatfile_926_http

LEGACY_CODE_TO_TYPE = {
    0: "formulario_afiliacion",
    1: "anexo_sedes",
    2: "listado_trabajadores",
    3: "comision",
    4: "carta",
    5: "camara_comercio",
    6: "cedula",
    7: "constancia_afiliacion",
    8: "rut",
    10: "entrega_documentos",
    11: "soporte_pagos",
    12: "contrato",
    13: "eps",
    14: "afp",
    15: "paz_y_salvo",
    16: "eps_afp",
    17: "detectar",
    19: "identificacion_peligros",
    20: "examen_preocupacional",
    21: "autorizacion_terceros",
    22: "historia_clinica",
    23: "afiliacion_eps",
    24: "afiliacion_afp",
    25: "carta_independiente_voluntario",
    26: "siarl",
    27: "beneficiario_final",
    28: "sat",
    98: "autorizacion",
    99: "imagen",
}

DOC_TYPE_LABELS = {
    "comision": "Comisión",
    "carta": "Cartas",
    "constancia_afiliacion": "Verificación",
    "cedula": "Cédula",
    "rut": "RUT",
    "camara_comercio": "Cámara de comercio",
    "contrato": "Contrato",
    "soporte_ingresos": "Pagos",
    "formulario_afiliacion": "Afiliación",
    "anexo_sedes": "Sedes",
    "listado_trabajadores": "Listados",
    "entrega_documentos": "Entrega Doc",
    "soporte_pagos": "Pagos",
    "identificacion_peligros": "Identificación de peligros",
    "examen_preocupacional": "Examen preocupacional",
    "beneficiario_final": "Beneficiario final",
    "autorizacion": "Autorización",
    "sat": "SAT",
}

LEGACY_INDEPENDIENTES_FIELDS = [
    "sr", "linea", "tipodocumento", "documento", "primer_apellido", "segundo_apellido", "primer_nombre",
    "segundo_nombre", "fecha_nacimiento", "sexo", "direccion", "departamento", "municipio", "zona", "localidad",
    "telefono", "celular", "correo", "eps", "codigo_eps", "afp", "codigo_afp", "arl_anterior",
    "codigo_arl_anterior", "tipo_cotizante", "subtipo_cotizante", "modalidad", "actividad_especial", "tipo_contrato",
    "transporte", "fecha_inicio_contrato", "fecha_fin_contrato", "meses_contrato", "valor_contrato", "valor_mensual",
    "ibc", "actividad_economica", "nombre_actividad", "clase_riesgo", "tasa_riesgo", "lunes", "martes", "miercoles",
    "jueves", "viernes", "sabado", "domingo", "d1", "d2", "d3", "d4", "d5", "d6", "d7", "d8", "d9", "d10",
    "d11", "d12", "d13", "d14", "d15", "d16", "d17", "d18", "d19", "d20", "d21", "d22", "d23", "d24",
    "codigo_ct", "nombre_ct", "actividad_economica_ct", "clase_riesgo_ct", "tasa_riesgo_ct", "direccion_ct",
    "departamento_ct", "ciudad_ct", "zona_ct", "telefono_ct", "celular_ct", "correo_ct", "localidad_ct", "lote",
    "tipo_salario",
]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _format_date_es(dt: datetime) -> str:
    months = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }
    return f"{dt.day} de {months.get(dt.month, str(dt.month))} de {dt.year}"


def get_cases_root() -> Path:
    root = Path(settings.cases_dir)
    root.mkdir(parents=True, exist_ok=True)
    return root


def get_document_registry_path() -> Path:
    path = Path(settings.document_registry_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def get_lote_counter_path() -> Path:
    path = Path(settings.lote_counter_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def get_case_dir(case_id: str) -> Path:
    return get_cases_root() / case_id


def get_case_metadata_path(case_id: str) -> Path:
    return get_case_dir(case_id) / "case.json"


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def normalize_haystack(value: Any) -> str:
    text = normalize_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip()


def fuzzy_text_score(query: str, candidate: Any) -> int:
    q = normalize_haystack(query)
    c = normalize_haystack(candidate)
    if not q or not c:
        return 0
    if q in c or c in q:
        return 12
    candidates = [c]
    candidates.extend(token for token in c.split() if len(token) >= 4)
    ratio = max(SequenceMatcher(None, q, piece).ratio() for piece in candidates)
    if ratio >= 0.9:
        return 12
    if ratio >= 0.8:
        return 8
    if ratio >= 0.72:
        return 4
    return 0


def slugify(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", value).strip("-").lower()
    return cleaned or "archivo"


def only_digits(value: Any) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def build_generated_lote_usuario(fecha_proceso: str) -> str:
    fecha_digits = only_digits(fecha_proceso)[:8] or datetime.now().strftime("%Y%m%d")
    path = get_lote_counter_path()
    state: Dict[str, Any] = {}
    if path.exists():
        try:
            state = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            state = {}
    last_date = only_digits(state.get("last_date"))[:8]
    last_seq = int(state.get("last_seq") or 0)
    next_seq = last_seq + 1 if last_date == fecha_digits else 1
    state = {
        "last_date": fecha_digits,
        "last_seq": next_seq,
        "updated_at": utc_now(),
    }
    path.write_text(json.dumps(state, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return f"{next_seq:012d}"


def _canonical_numeric_value(value: Any) -> str:
    digits = only_digits(value)
    stripped = digits.lstrip("0")
    return stripped or digits


def _numeric_similarity_match(expected: str, candidate: str) -> bool:
    expected_digits = _canonical_numeric_value(expected)
    candidate_digits = _canonical_numeric_value(candidate)
    if not expected_digits or not candidate_digits:
        return False
    if expected_digits == candidate_digits:
        return True
    if abs(len(expected_digits) - len(candidate_digits)) > 1:
        return False
    if expected_digits.startswith(candidate_digits) or candidate_digits.startswith(expected_digits):
        return True
    mismatches = 0
    i = 0
    j = 0
    while i < len(expected_digits) and j < len(candidate_digits):
        if expected_digits[i] == candidate_digits[j]:
            i += 1
            j += 1
            continue
        mismatches += 1
        if mismatches > 1:
            return False
        if len(expected_digits) > len(candidate_digits):
            i += 1
        elif len(candidate_digits) > len(expected_digits):
            j += 1
        else:
            i += 1
            j += 1
    if i < len(expected_digits) or j < len(candidate_digits):
        mismatches += 1
    return mismatches <= 1


def _best_numeric_candidate(expected: str, values: List[Any]) -> str:
    expected_digits = _canonical_numeric_value(expected)
    candidates: List[str] = []
    for raw in values:
        digits = _canonical_numeric_value(raw)
        if digits:
            candidates.append(digits)
    exact = [candidate for candidate in candidates if candidate == expected_digits]
    if exact:
        return exact[0]
    close = [candidate for candidate in candidates if _numeric_similarity_match(expected_digits, candidate)]
    if close:
        close.sort(key=lambda item: (abs(len(item) - len(expected_digits)), -len(item)))
        return close[0]
    return ""


def _normalize_company_nit(value: Any, docs: Optional[List[Dict[str, Any]]] = None) -> str:
    digits = only_digits(value)
    if not digits:
        return ""
    if len(digits) != 10 or not docs:
        return digits
    prefix = digits[:-1]
    doc_nits = set()
    for doc in docs:
        fields = doc.get("fields") or {}
        doc_nit = only_digits(fields.get("nit", ""))
        if 8 <= len(doc_nit) <= 10:
            doc_nits.add(doc_nit)
    if prefix in doc_nits:
        return prefix
    return digits


def _extract_name_from_cedula_text(text: str) -> str:
    normalized = normalize_text(text).upper()
    direct_patterns = [
        r"([A-ZÁÉÍÓÚÑ ]{4,80})\s+APELLIDOS\b[^A-ZÁÉÍÓÚÑ]{0,10}([A-ZÁÉÍÓÚÑ .]{4,80})\s+NOMBRES",
        r"APELLIDOS\b[^A-ZÁÉÍÓÚÑ]{0,10}([A-ZÁÉÍÓÚÑ ]{4,80})\s+NOMBRES\b[^A-ZÁÉÍÓÚÑ]{0,10}([A-ZÁÉÍÓÚÑ ]{4,80})",
    ]
    for pattern in direct_patterns:
        match = re.search(pattern, normalized, flags=re.IGNORECASE)
        if match:
            groups = [normalize_text(group).strip(" .,-") for group in match.groups() if normalize_text(group)]
            if len(groups) == 2:
                return normalize_text(f"{groups[1]} {groups[0]}")

    compact = re.sub(r"[^A-ZÁÉÍÓÚÑ ]+", " ", normalized)
    compact = re.sub(r"\s+", " ", compact).strip()
    match = re.search(r"([A-ZÁÉÍÓÚÑ ]{4,80}) APELLIDOS ([A-ZÁÉÍÓÚÑ ]{4,80}) NOMBRES", compact)
    if match:
        surname = normalize_text(match.group(1))
        names = normalize_text(match.group(2))
        return normalize_text(f"{names} {surname}")

    if "APELLIDOS" in compact and "NOMBRES" in compact:
        before_apellidos, after_apellidos = compact.split("APELLIDOS", 1)
        names_segment, _, _ = after_apellidos.partition("NOMBRES")
        surname_tokens = before_apellidos.split()[-4:]
        name_tokens = names_segment.split()[:4]
        surname = normalize_text(" ".join(token for token in surname_tokens if len(token) > 1))
        names = normalize_text(" ".join(token for token in name_tokens if len(token) > 1))
        if surname and names:
            return normalize_text(f"{names} {surname}")
    return ""


def load_case(case_id: str) -> Dict[str, Any]:
    metadata_path = get_case_metadata_path(case_id)
    if not metadata_path.exists():
        raise FileNotFoundError(case_id)
    return json.loads(metadata_path.read_text(encoding="utf-8"))


def save_case(case_payload: Dict[str, Any]) -> Dict[str, Any]:
    case_id = str(case_payload["id"])
    case_dir = get_case_dir(case_id)
    case_dir.mkdir(parents=True, exist_ok=True)
    get_case_metadata_path(case_id).write_text(
        json.dumps(case_payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return case_payload


def list_cases() -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for path in sorted(get_cases_root().glob("*/case.json"), reverse=True):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            rows.append(payload)
        except Exception:
            continue
    return rows


def get_case_file_path(case_id: str, filename: str) -> Path:
    case_dir = get_case_dir(case_id) / "files"
    target = case_dir / Path(filename).name
    if not target.exists():
        raise FileNotFoundError(filename)
    return target


def search_cases(query: str, limit: int = 10) -> List[Dict[str, Any]]:
    needle = normalize_haystack(query)
    query_tokens: List[str] = []
    for token in needle.split():
        if len(token) < 3:
            continue
        query_tokens.append(token)
        if token.endswith("s") and len(token) >= 5:
            query_tokens.append(token[:-1])
    results: List[Dict[str, Any]] = []
    for payload in list_cases():
        analysis = payload.get("analysis") or {}
        profile = (analysis.get("xlsx_profile") or {}).get("profile") or {}
        docs = analysis.get("documents") or []
        case_haystack = " ".join(
            [
                payload.get("label", ""),
                payload.get("lote_usuario", ""),
                profile.get("empresa", ""),
                profile.get("nombre", ""),
                profile.get("documento", ""),
                profile.get("nit", ""),
                profile.get("tipo_afiliado", ""),
                profile.get("lote_usuario", ""),
            ]
        )
        case_norm = normalize_haystack(case_haystack)
        score = 0
        if needle and needle in case_norm:
            score += 10
        score += fuzzy_text_score(query, profile.get("empresa", ""))
        score += fuzzy_text_score(query, payload.get("label", "")) // 2
        score += fuzzy_text_score(query, profile.get("nombre", "")) // 2
        for token in query_tokens:
            if token in case_norm:
                score += 4
        if profile.get("empresa"):
            score += 3
        if profile.get("nombre"):
            score += 2
        if profile.get("documento"):
            score += 2
        if profile.get("nit"):
            score += 2
        if not profile.get("empresa"):
            score -= 6
        if not _looks_like_person_name(profile.get("nombre", "")):
            score -= 3
        if profile.get("empresa") and not _looks_like_company_name(profile.get("empresa", "")):
            score -= 14
        if len(docs) < 2:
            score -= 18
        if not docs and not profile.get("empresa") and not profile.get("documento"):
            score -= 25
        if not profile.get("empresa") and not profile.get("nombre"):
            score -= 20
        if payload.get("status") != "analyzed":
            score -= 10
        if normalize_haystack(payload.get("label", "")).startswith("case-"):
            score -= 1
        updated_at = str(payload.get("updated_at") or "")
        if updated_at.startswith("2026-03-16T13:"):
            score += 2
        matched_documents: List[Dict[str, Any]] = []
        entity_document_hint = only_digits(profile.get("documento", ""))
        for doc in docs:
            fields = doc.get("fields") or {}
            preview = str(doc.get("text_preview") or "")
            raw_doc_person_name = _normalize_person_name(
                fields.get("representative_name", "") or _extract_name_from_cedula_text(preview)
            )
            doc_person_name = raw_doc_person_name if _looks_like_person_name(raw_doc_person_name) else ""
            if not entity_document_hint and doc.get("document_type") in {"cedula", "formulario_afiliacion", "anexo_sedes", "carta", "constancia_afiliacion"}:
                entity_document_hint = only_digits(fields.get("representative_document", "") or fields.get("document_number", ""))
            doc_haystack = normalize_haystack(
                " ".join(
                    [
                        doc.get("filename", ""),
                        doc.get("document_type", ""),
                        preview,
                        str(fields.get("company_name") or ""),
                        doc_person_name,
                        str(fields.get("document_number") or ""),
                        str(fields.get("nit") or ""),
                    ]
                )
            )
            doc_score = 0
            if needle and needle in doc_haystack:
                doc_score += 8
            for token in query_tokens:
                if token in doc_haystack:
                    doc_score += 3
            if doc_score:
                matched_documents.append(
                    {
                        "filename": doc.get("filename"),
                        "document_type": doc.get("document_type"),
                        "person_name": doc_person_name,
                        "score": doc_score,
                        "snippet": preview[:280],
                    }
                )
                score += doc_score
        if score <= 0:
            continue
        matched_documents.sort(key=lambda item: item["score"], reverse=True)
        results.append(
            {
                "case_id": payload.get("id"),
                "label": payload.get("label"),
                "status": payload.get("status"),
                "score": score,
                "updated_at": payload.get("updated_at") or "",
                "profile": profile,
                "lote_usuario": payload.get("lote_usuario") or profile.get("lote_usuario") or "",
                "entity_document_hint": entity_document_hint,
                "doc_signature": tuple(
                    sorted(
                        f"{normalize_haystack(doc.get('filename', ''))}|{normalize_haystack(doc.get('document_type', ''))}"
                        for doc in docs
                    )
                ),
                "matched_documents": matched_documents[:8],
                "received_summary": ((analysis.get("checklist") or {}).get("received_summary") or [])[:20],
                "decision": analysis.get("decision") or {},
                "precheck": ((analysis.get("validacion_resumen") or {}).get("precheck") or {}),
                "executive_report": analysis.get("reporte_ejecutivo") or {},
            }
        )
    def _profile_quality(item: Dict[str, Any]) -> int:
        profile = item.get("profile") or {}
        quality = 0
        if _looks_like_company_name(profile.get("empresa", "")):
            quality += 6
        elif normalize_text(profile.get("empresa", "")):
            quality += 2
        if _looks_like_person_name(profile.get("nombre", "")):
            quality += 3
        if only_digits(profile.get("nit", "")):
            quality += 3
        if only_digits(profile.get("documento", "")) or only_digits(item.get("entity_document_hint", "")):
            quality += 3
        if item.get("received_summary"):
            quality += 2
        if item.get("matched_documents"):
            quality += 1
        if not normalize_text(profile.get("empresa", "")):
            quality -= 8
        if not only_digits(profile.get("nit", "")) and not only_digits(profile.get("documento", "")):
            quality -= 4
        return quality

    results.sort(
        key=lambda item: (
            int(item.get("score") or 0),
            _profile_quality(item),
            1 if (item.get("precheck") or {}).get("approved") else 0,
            1 if ((item.get("decision") or {}).get("recommended_status") == "aprobable") else 0,
            str(item.get("updated_at") or ""),
        ),
        reverse=True,
    )
    deduped: List[Dict[str, Any]] = []
    seen = set()
    seen_doc_signatures = set()
    seen_documents = set()
    for item in results:
        doc_signature = tuple(item.get("doc_signature") or ())
        dedupe_key = (
            only_digits((item.get("profile") or {}).get("nit") or ""),
            only_digits((item.get("profile") or {}).get("documento") or item.get("entity_document_hint") or ""),
        )
        document_key = normalize_haystack((item.get("profile") or {}).get("documento") or item.get("entity_document_hint") or "")
        company_value = (item.get("profile") or {}).get("empresa") or ""
        name_value = (item.get("profile") or {}).get("nombre") or ""
        strong_profile = _looks_like_company_name(company_value) and _looks_like_person_name(name_value)
        weak_profile = not strong_profile
        if doc_signature and doc_signature in seen_doc_signatures and weak_profile:
            continue
        if document_key and document_key in seen_documents and weak_profile:
            continue
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        if doc_signature:
            seen_doc_signatures.add(doc_signature)
        if document_key:
            seen_documents.add(document_key)
        deduped.append(item)
        if len(deduped) >= limit:
            break
    return deduped


def rebuild_document_registry() -> Dict[str, Any]:
    registry: List[Dict[str, Any]] = []
    for payload in list_cases():
        analysis = payload.get("analysis") or {}
        profile = (analysis.get("xlsx_profile") or {}).get("profile") or {}
        file_map = {item.get("filename"): item for item in payload.get("files") or []}
        for doc in analysis.get("documents") or []:
            fields = doc.get("fields") or {}
            filename = doc.get("filename") or ""
            file_entry = file_map.get(filename) or {}
            ocr_text = normalize_text(doc.get("ocr_text", "") or doc.get("text_preview", ""))
            doc_type = str(doc.get("document_type") or "")
            candidate_person_name = _normalize_person_name(
                fields.get("representative_name", "") or _extract_name_from_cedula_text(ocr_text)
            )
            person_name_source = "document"
            if not _looks_like_person_name(candidate_person_name) and doc_type not in {"entrega_documentos"}:
                candidate_person_name = _normalize_person_name(profile.get("nombre", ""))
                person_name_source = "profile"
            derived_person_name = candidate_person_name if _looks_like_person_name(candidate_person_name) else ""
            explicit_person_document = only_digits(fields.get("representative_document", "") or fields.get("document_number", ""))
            fallback_person_document = only_digits(profile.get("documento", ""))
            identity_document_types = {"cedula", "formulario_afiliacion", "anexo_sedes", "carta", "constancia_afiliacion"}
            if (
                doc_type == "cedula"
                and fallback_person_document
                and explicit_person_document
                and (
                    len(explicit_person_document) < 7
                    or explicit_person_document.startswith("0500")
                    or not _numeric_similarity_match(explicit_person_document, fallback_person_document)
                )
            ):
                explicit_person_document = fallback_person_document
            if explicit_person_document:
                derived_person_document = explicit_person_document
                person_document_source = "document"
            elif doc_type in identity_document_types and fallback_person_document:
                derived_person_document = fallback_person_document
                person_document_source = "profile"
            else:
                derived_person_document = ""
                person_document_source = ""
            registry.append(
                {
                    "case_id": payload.get("id"),
                    "label": payload.get("label"),
                    "status": payload.get("status"),
                    "filename": filename,
                    "stored_path": file_entry.get("stored_path", ""),
                    "document_type": doc_type,
                    "legacy_code": doc.get("legacy_code"),
                    "used_ocr": bool(doc.get("used_ocr")),
                    "company": _clean_company_name(profile.get("empresa", "") or fields.get("company_name", "")),
                    "nit": only_digits(profile.get("nit", "") or fields.get("nit", "")),
                    "person_name": derived_person_name,
                    "person_name_source": person_name_source if derived_person_name else "",
                    "person_document": derived_person_document,
                    "person_document_source": person_document_source,
                    "document_number": only_digits(fields.get("document_number", "")),
                    "ocr_text": ocr_text,
                    "text_preview": normalize_text(doc.get("text_preview", "")),
                }
            )
    path = get_document_registry_path()
    preferred_by_doc: Dict[tuple, int] = {}
    filtered_registry: List[Dict[str, Any]] = []
    for entry in registry:
        key = (
            normalize_haystack(entry.get("filename", "")),
            normalize_haystack(entry.get("document_type", "")),
            only_digits(entry.get("person_document", "")) or only_digits(entry.get("document_number", "")),
        )
        quality = 0
        if entry.get("company") and _looks_like_company_name(entry.get("company", "")):
            quality += 10
        if entry.get("person_name") and _looks_like_person_name(entry.get("person_name", "")):
            quality += 6
        if only_digits(entry.get("person_document", "")):
            quality += 4
        if str(entry.get("label", "")).startswith("case-"):
            quality += 1
        previous = preferred_by_doc.get(key)
        if previous is None or quality > previous:
            preferred_by_doc[key] = quality
    for entry in registry:
        key = (
            normalize_haystack(entry.get("filename", "")),
            normalize_haystack(entry.get("document_type", "")),
            only_digits(entry.get("person_document", "")) or only_digits(entry.get("document_number", "")),
        )
        quality = 0
        if entry.get("company") and _looks_like_company_name(entry.get("company", "")):
            quality += 10
        if entry.get("person_name") and _looks_like_person_name(entry.get("person_name", "")):
            quality += 6
        if only_digits(entry.get("person_document", "")):
            quality += 4
        if str(entry.get("label", "")).startswith("case-"):
            quality += 1
        if quality < preferred_by_doc.get(key, quality):
            continue
        filtered_registry.append(entry)
    path.write_text(json.dumps(filtered_registry, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return {"documents": len(filtered_registry), "documents_raw": len(registry), "path": str(path)}


def load_document_registry() -> List[Dict[str, Any]]:
    path = get_document_registry_path()
    if not path.exists():
        rebuild_document_registry()
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return []


def search_document_registry(query: str, limit: int = 12) -> List[Dict[str, Any]]:
    needle = normalize_haystack(query)
    query_digits = only_digits(query)
    generic_query_terms = {
        "cedula",
        "cédula",
        "documento",
        "documentos",
        "representante",
        "representangte",
        "representate",
        "representaante",
        "rut",
        "camara",
        "cámara",
        "formulario",
        "pdf",
        "soporte",
        "sede",
        "entrega",
    }
    tokens: List[str] = []
    for token in needle.split():
        if len(token) < 3:
            continue
        tokens.append(token)
        if token.endswith("s") and len(token) >= 5:
            tokens.append(token[:-1])
    meaningful_tokens = [token for token in tokens if token not in generic_query_terms]
    results: List[Dict[str, Any]] = []
    for item in load_document_registry():
        ocr_text = normalize_haystack(item.get("ocr_text", ""))
        company = normalize_haystack(item.get("company", ""))
        person_name = normalize_haystack(item.get("person_name", ""))
        person_document = only_digits(item.get("person_document", ""))
        document_number = only_digits(item.get("document_number", ""))
        document_type = normalize_haystack(item.get("document_type", ""))
        label = normalize_haystack(item.get("label", ""))
        person_name_source = str(item.get("person_name_source") or "")
        person_document_source = str(item.get("person_document_source") or "")
        haystack = " ".join([label, company, person_name, item.get("nit", ""), person_document, document_number, document_type, ocr_text]).strip()
        score = 0
        if needle and needle in haystack:
            score += 12
        for token in tokens:
            if token in haystack:
                score += 4
            if token and token in company:
                score += 3
            if token and token in person_name:
                score += 4
            if token and token in document_type:
                score += 3
            if token and token in ocr_text:
                score += 1
        for token in meaningful_tokens:
            if token in person_name:
                score += 10
            elif token in company:
                score += 8
            elif token in haystack:
                score += 5
            else:
                score -= 10
        if query_digits:
            if query_digits == person_document:
                score += 32
            elif query_digits == document_number:
                score += 24
            elif _numeric_similarity_match(query_digits, person_document):
                score += 20
            elif _numeric_similarity_match(query_digits, document_number):
                score += 14
        if item.get("document_type") == "cedula" and query_digits and (query_digits == person_document or _numeric_similarity_match(query_digits, person_document)):
            score += 10
        if item.get("document_type") == "cedula" and ("cedula de ciudadania" in ocr_text or "identificacion personal" in ocr_text):
            score += 18
        if item.get("document_type") == "camara_comercio" and "camara" in needle:
            score += 8
        if item.get("document_type") == "rut" and "rut" in needle:
            score += 8
        if item.get("document_type") == "entrega_documentos":
            score -= 6
        if person_document_source == "profile" and item.get("document_type") not in {"cedula", "formulario_afiliacion", "anexo_sedes", "carta", "constancia_afiliacion"}:
            score -= 10
        if person_name_source == "profile" and item.get("document_type") in {"entrega_documentos", "rut", "camara_comercio", "pdf"}:
            score -= 5
        if item.get("company"):
            score += 2
        else:
            score -= 10
        if item.get("company") and not _looks_like_company_name(item.get("company", "")):
            score -= 12
        if item.get("person_name"):
            score += 1
        else:
            score -= 2
        label_norm = normalize_haystack(item.get("label", ""))
        company_norm = normalize_haystack(item.get("company", ""))
        if "demo" in label_norm or "demo" in company_norm:
            score -= 8
        if label_norm.startswith("case-"):
            score -= 2
        if (not item.get("company")) and label_norm:
            score -= 10
        if item.get("document_type") == "formulario_afiliacion" and not item.get("company"):
            score -= 8
        if item.get("filename", "").lower().endswith(".txt"):
            score -= 4
        if item.get("used_ocr"):
            score += 1
        if score <= 0:
            continue
        enriched = dict(item)
        enriched["score"] = score
        results.append(enriched)
    results.sort(key=lambda item: item["score"], reverse=True)
    deduped: List[Dict[str, Any]] = []
    seen = set()
    for item in results:
        logical_filename = re.sub(r"(-\d+)(\.[a-z0-9]+)$", r"\2", Path(item.get("filename", "")).name.lower())
        dedupe_key = (
            str(item.get("case_id") or ""),
            only_digits(item.get("person_document", "")) or only_digits(item.get("document_number", "")),
            normalize_haystack(item.get("company", "")),
            normalize_haystack(item.get("document_type", "")),
            logical_filename,
        )
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        deduped.append(item)
        if len(deduped) >= limit:
            break
    return deduped


def create_case_record(label: str, source_files: List[Dict[str, Any]]) -> Dict[str, Any]:
    case_id = f"case-{uuid.uuid4().hex[:10]}"
    payload = {
        "id": case_id,
        "label": normalize_text(label) or case_id,
        "status": "uploaded",
        "created_at": utc_now(),
        "updated_at": utc_now(),
        "files": source_files,
        "analysis": None,
    }
    return save_case(payload)


def _clean_company_name(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = re.split(r"\b(?:nit|cc|numero|número|cps-f)\b", text, maxsplit=1, flags=re.IGNORECASE)[0]
    text = re.sub(r"\s+", " ", text).strip(" .,-")
    return text


def _normalize_company_compare(value: str) -> str:
    text = normalize_haystack(_clean_company_name(value))
    text = re.sub(r"[^a-z0-9]+", "", text)
    text = text.replace("sas", "sas")
    return text


def _looks_like_person_name(value: str) -> bool:
    text = normalize_text(value)
    if len(text) < 8:
        return False
    lowered = text.lower()
    blocked = [
        "en mora",
        "acuerdo de pago",
        "urbana",
        "codigo",
        "formulario",
        "número de documento",
        "numero de documento",
        "representante legal de colmena",
        "de participación",
        "de documento",
    ]
    if any(token in lowered for token in blocked):
        return False
    tokens = [token for token in text.split() if token]
    return len(tokens) >= 2


def _normalize_person_name(value: str) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = re.sub(r"\bSEG\s+NDO\b", "SEGUNDO", text, flags=re.IGNORECASE)
    text = re.sub(r"\bSEG\s+NDA\b", "SEGUNDA", text, flags=re.IGNORECASE)
    text = re.sub(r"\bPOLICARPO\s+SEG\s+NDO\b", "POLICARPO SEGUNDO", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip(" .,-")
    return text


def _looks_like_company_name(value: str) -> bool:
    text = _clean_company_name(value)
    if len(text) < 5:
        return False
    lowered = text.lower()
    blocked = [
        "cedula de ciudadania",
        "del empleador",
        "numero de documento",
        "número de documento",
        "representante legal",
        "fecha inicio",
        "estado de cuenta",
        "lugar de nacimiento",
        "fecha y lugar",
    ]
    if any(token in lowered for token in blocked):
        return False
    return True


def _looks_like_cedula_document(name_txt: str, haystack: str) -> bool:
    positive_markers = [
        "cedula de ciudadania",
        "identificacion personal",
        "registraduria nacional",
        "apellidos",
        "nombres",
        "lugar de nacimiento",
        "fecha y lugar de expedicion",
        "sexo",
    ]
    negative_markers = [
        "señores arl",
        "representante legal de la empresa",
        "constancia de afiliacion",
        "direccion de aseguramiento",
        "numero de identificacion tributaria",
        "tarjeta profesional",
        "revisor fiscal",
        "camara de comercio",
        "certificado de existencia",
        "formulario de afiliacion",
        "riesgos laborales",
        "planilla resumen",
        "resumen general de pago en linea",
        "resumen general de pago en línea",
        "estado de cuenta",
        "beneficiario final",
        "beneficiarios finales",
        "participacion directa o indirecta mayor al 5",
        "participación directa o indirecta mayor al 5",
        "informacion de la compania",
        "información de la compañía",
        "conformacion de la sociedad",
        "conformación de la sociedad",
        "accionista",
    ]
    positive_hits = sum(1 for marker in positive_markers if marker in haystack)
    fuzzy_positive_patterns = [
        r"registrad\w+\s+nacional",
        r"indice\s+derech\w*",
        r"fecha\s+y\s+lugar\s+de\s+expedicion",
        r"fecha\s+y\s+lug\w*\s+expedicion",
        r"lugar\s+de\s+nac\w+",
        r"\brh\b",
        r"\bestatura\b",
        r"lugar\s+de\s+nacimiento",
    ]
    negative_hits = sum(1 for marker in negative_markers if marker in haystack)
    positive_hits += sum(1 for pattern in fuzzy_positive_patterns if re.search(pattern, haystack))
    if any(token in name_txt for token in ["cedula", "cc_"]):
        positive_hits += 2
    if "cedula de ciudadania" in haystack or "identificacion personal" in haystack:
        positive_hits += 2
    if "fecha y lug" in haystack and "expedicion" in haystack:
        positive_hits += 2
    if "lugar de nac" in haystack:
        positive_hits += 2
    if re.search(r"\b\d{7,10}\b", haystack):
        positive_hits += 1
    if negative_hits >= 2 and positive_hits < 5:
        return False
    return positive_hits >= 4 and negative_hits <= 1


def _looks_like_camara_document(haystack: str) -> bool:
    positive_markers = [
        "camara de comercio",
        "certificado de existencia y representacion legal",
        "matricula mercantil",
        "codigo de verificacion",
        "fecha expedicion",
        "sede virtual",
        "certificadoselectronicos",
        "recibo no.",
    ]
    negative_markers = [
        "señores arl",
        "senores arl",
        "respetados señores",
        "respetados senores",
        "por medio de la presente",
        "constancia de afiliacion",
        "planilla resumen",
        "estado de cuenta",
        "autorizacion de tratamiento",
    ]
    positive_hits = sum(1 for marker in positive_markers if marker in haystack)
    negative_hits = sum(1 for marker in negative_markers if marker in haystack)
    return positive_hits >= 2 and negative_hits == 0


def _looks_like_entrega_documentos(haystack: str) -> bool:
    positive_markers = [
        "comprobante entrega de documentos",
        "pago de reconocimiento variable integral",
        "documentos anexos a la afiliacion",
        "datos de la afiliacion",
        "nro. de contrato",
    ]
    return sum(1 for marker in positive_markers if marker in haystack) >= 2


def _looks_like_carta_document(haystack: str) -> bool:
    positive_markers = [
        "señores arl",
        "senores arl",
        "respetados señores",
        "respetados senores",
        "por medio de la presente",
        "representante legal de la empresa",
        "me permito informarle",
        "solicitud de desafiliacion",
        "solicitud de desafiliación",
        "se adjuntan los siguientes documentos",
        "agradeciendo su colaboracion",
        "agradeciendo su colaboración",
        "piensalo bien",
        "piénsalo bien",
        "servicioalcliente@positiva.com.co",
        "se recibe su solicitud de traslado",
    ]
    return sum(1 for marker in positive_markers if marker in haystack) >= 2


def _looks_like_constancia_afiliacion(haystack: str) -> bool:
    positive_markers = [
        "constancia de afiliacion",
        "hace constar",
        "direccion de aseguramiento",
        "estado numero de centros de trabajo",
        "codigo unico de generacion",
        "reporte consulta no.",
        "listas propias",
        "listas informativas",
        "risk consulting global group certifica",
        "consulta de procesos en la rama judicial",
        "codigo de radicado asignado",
        "peticion a arl ha sido enviada",
        "peticion a arl",
        "transaccion financiera exitosa",
        "transaccion exitosa",
        "ley de transparencia",
    ]
    hits = sum(1 for marker in positive_markers if marker in haystack)
    return hits >= 2 or any(
        marker in haystack
        for marker in [
            "codigo de radicado asignado",
            "reporte consulta no.",
            "consulta de procesos en la rama judicial",
            "risk consulting global group certifica",
            "ley de transparencia",
        ]
    )


def _looks_like_rut_document(haystack: str) -> bool:
    positive_markers = [
        "registro unico tributario",
        "registro único tributario",
        "direccion de impuestos y aduanas",
        "dirección de impuestos y aduanas",
        "espacio reservado para la dian",
        "numero de formulario",
        "número de formulario",
        "direccion seccional",
        "dirección seccional",
        "buzon electronico",
        "buzón electrónico",
        "numero de identificacion tributaria",
        "número de identificación tributaria",
        "95. numero de identificacion tributaria",
        "95. número de identificación tributaria",
        "matriz o controlante",
        "espacio reservado para la dian",
        "entidades o institutos de derecho publico",
        "entidades o institutos de derecho público",
        "numero de formulario",
        "número de formulario",
    ]
    negative_markers = [
        "outlook",
        "desafiliacion por traslado",
        "desafiliación por traslado",
        "traslado voluntario de arl",
        "se recibe su solicitud de traslado",
        "piensalo bien",
        "positiva.gov.co",
        "hasta el viernes",
        "correo electronico",
        "correo electrónico",
        "solicitud de desafiliacion",
        "solicitud de desafiliación",
        "se adjuntan los siguientes documentos",
        "agradeciendo su colaboracion",
        "agradeciendo su colaboración",
        "se recibe su solicitud de traslado",
    ]
    hits = sum(1 for marker in positive_markers if marker in haystack)
    negative_hits = sum(1 for marker in negative_markers if marker in haystack)
    return hits >= 2 and negative_hits == 0


def _looks_like_autorizacion_document(haystack: str) -> bool:
    positive_markers = [
        "autorizacion de tratamiento de datos",
        "autorización de tratamiento de datos",
        "tratamiento de los datos personales",
        "tratamiento de datos personales",
        "acuerdo de transmision de datos personales",
        "acuerdo de transmisión de datos personales",
        "finalidades que han sido autorizadas",
        "responsable sobre la informacion",
        "responsable sobre la información",
        "encargado",
        "clausula sexta",
        "cláusula sexta",
        "clausula septima",
        "cláusula séptima",
        "remitir a traves de los canales",
        "remitir a través de los canales",
        "declaro que he sido informado que",
        "recopilar, analizar, consultar, validar y procesar",
        "fundacion grupo social",
        "fundación grupo social",
    ]
    return sum(1 for marker in positive_markers if marker in haystack) >= 2


def _looks_like_comision_document(haystack: str) -> bool:
    positive_markers = [
        "labor de intermediacion en riesgos laborales",
        "labor de intermediación en riesgos laborales",
        "ha designado a",
        "intermediacion en el ramo de riesgos laborales es voluntaria",
        "intermediación en el ramo de riesgos laborales es voluntaria",
        "paragrafo 5 del articulo 11 de la ley 1562 de 2012",
        "parágrafo 5 del artículo 11 de la ley 1562 de 2012",
    ]
    return sum(1 for marker in positive_markers if marker in haystack) >= 2


def _looks_like_beneficiario_final_document(haystack: str) -> bool:
    positive_markers = [
        "beneficiario final",
        "beneficiarios finales",
        "registro unico de beneficiarios finales",
        "registro único de beneficiarios finales",
        "participacion directa o indirecta mayor al 5",
        "participación directa o indirecta mayor al 5",
        "informacion de la compania",
        "información de la compañía",
        "conformacion de la sociedad",
        "conformación de la sociedad",
        "accionista",
        "representante legaladministrador",
    ]
    return sum(1 for marker in positive_markers if marker in haystack) >= 2


def _extract_page_number(filename: str) -> int:
    match = re.search(r"__p(\d+)\.pdf$", str(filename or ""), flags=re.IGNORECASE)
    return int(match.group(1)) if match else -1


def _apply_document_classification_overrides(docs: List[Dict[str, Any]]) -> None:
    prefix_groups: Dict[str, List[Dict[str, Any]]] = {}
    for doc in docs:
        filename = str(doc.get("filename") or "")
        prefix = filename.split("__p", 1)[0]
        prefix_groups.setdefault(prefix, []).append(doc)

    for group in prefix_groups.values():
        group.sort(key=lambda item: _extract_page_number(str(item.get("filename") or "")))

        for index, doc in enumerate(group):
            haystack = normalize_haystack(
                f"{doc.get('filename', '')} {doc.get('ocr_text', '') or doc.get('text_preview', '')}"
            )

            if _looks_like_beneficiario_final_document(haystack):
                doc["document_type"] = "beneficiario_final"
                doc["legacy_code"] = 27
                doc["legacy_label"] = LEGACY_CODE_TO_TYPE.get(27, "")
                doc["code_source"] = "post_beneficiario_precise"
                continue

            if doc.get("document_type") == "rut" and _looks_like_carta_document(haystack):
                doc["document_type"] = "carta"
                doc["legacy_code"] = 4
                doc["legacy_label"] = LEGACY_CODE_TO_TYPE.get(4, "")
                doc["code_source"] = "post_carta_from_rut"
                continue

            if doc.get("document_type") == "pdf":
                if (
                    ("lugar de nac" in haystack or re.search(r"lugar\s+de\s+n[a-z]{2,}", haystack))
                    and ("fecha y lug" in haystack or "expedicion" in haystack)
                ):
                    doc["document_type"] = "cedula"
                    doc["legacy_code"] = 6
                    doc["legacy_label"] = LEGACY_CODE_TO_TYPE.get(6, "")
                    doc["code_source"] = "post_cedula_identity_layout"
                    continue

                if _looks_like_cedula_document(normalize_haystack(doc.get("filename", "")), haystack):
                    doc["document_type"] = "cedula"
                    doc["legacy_code"] = 6
                    doc["legacy_label"] = LEGACY_CODE_TO_TYPE.get(6, "")
                    doc["code_source"] = "post_cedula_from_pdf"
                    continue

                if any(
                    marker in haystack
                    for marker in [
                        "resumen general de pago en linea",
                        "resumen general de pago en línea",
                        "resumen general de pago en inea",
                        "aportes resumen general de pago en linea",
                        "aportes resumen general de pago en línea",
                        "aportes resumen general de pago en inea",
                    ]
                ):
                    doc["document_type"] = "soporte_ingresos"
                    doc["legacy_code"] = 11
                    doc["legacy_label"] = LEGACY_CODE_TO_TYPE.get(11, "")
                    doc["code_source"] = "post_planilla_resumen"
                    continue

                prev_doc = group[index - 1] if index > 0 else None
                next_doc = group[index + 1] if index + 1 < len(group) else None
                text_preview = (doc.get("text_preview") or "").strip()
                if (
                    not text_preview
                    and prev_doc
                    and next_doc
                    and prev_doc.get("document_type") == "rut"
                    and next_doc.get("document_type") == "soporte_ingresos"
                ):
                    doc["document_type"] = "cedula"
                    doc["legacy_code"] = 6
                    doc["legacy_label"] = LEGACY_CODE_TO_TYPE.get(6, "")
                    doc["code_source"] = "post_blank_cedula_between_rut_and_pagos"


def _classify_document(filename: str, text: str) -> Dict[str, Any]:
    name_txt = normalize_haystack(filename)
    text_txt = normalize_haystack(text)
    haystack = f"{name_txt} {text_txt}".strip()
    lower_name = str(filename or "").lower()

    if "formulario de afiliacion" in name_txt:
        return {"document_type": "formulario_afiliacion", "legacy_code": 0, "code_source": "name_override_formulario"}
    if re.search(r"\bsede[\s._-]*\d+\b", name_txt) or re.search(r"\bsedes?\b", name_txt):
        return {"document_type": "anexo_sedes", "legacy_code": 1, "code_source": "name_override_sede"}
    if re.match(r"^sede\d+(?:__p\d+)?\.pdf$", lower_name):
        return {"document_type": "anexo_sedes", "legacy_code": 1, "code_source": "name_override_sede_compact"}
    if (
        "a. afiliacion" in haystack
        or "a. afiliación" in haystack
        or ("b. traslado" in haystack and "c. terminacion" in haystack)
        or ("cps-f-216" in haystack)
    ):
        return {"document_type": "formulario_afiliacion", "legacy_code": 0, "code_source": "ocr_formulario_precise"}
    if _looks_like_beneficiario_final_document(haystack):
        return {"document_type": "beneficiario_final", "legacy_code": 27, "code_source": "ocr_beneficiario_precise"}
    if _looks_like_cedula_document(name_txt, haystack):
        return {"document_type": "cedula", "legacy_code": 6, "code_source": "ocr_cedula_precise"}
    if _looks_like_rut_document(haystack):
        return {"document_type": "rut", "legacy_code": 8, "code_source": "ocr_rut_precise"}
    if _looks_like_entrega_documentos(haystack):
        return {"document_type": "entrega_documentos", "legacy_code": 10, "code_source": "ocr_entrega_precise"}
    if _looks_like_camara_document(haystack):
        return {"document_type": "camara_comercio", "legacy_code": 5, "code_source": "ocr_camara_precise"}
    if _looks_like_constancia_afiliacion(haystack):
        return {"document_type": "constancia_afiliacion", "legacy_code": 7, "code_source": "ocr_constancia"}
    if _looks_like_autorizacion_document(haystack):
        return {"document_type": "autorizacion", "legacy_code": 98, "code_source": "ocr_autorizacion_precise"}
    if _looks_like_comision_document(haystack):
        return {"document_type": "comision", "legacy_code": 3, "code_source": "ocr_comision_precise"}
    if _looks_like_carta_document(haystack):
        return {"document_type": "carta", "legacy_code": 4, "code_source": "ocr_carta"}
    planilla_markers = [
        "aportes planilla resumen",
        "planilla resumen en linea",
        "planilla resumen en línea",
        "resumen general de pago en inea",
        "planilla resumen",
        "resumen general de pago en linea",
        "resumen general de pago en línea",
        "aportes resumen general de pago en linea",
        "aportes resumen general de pago en línea",
        "aportes resumen general de pago en inea",
        "resumen de pago riesgo",
        "datos generales del aportante",
        "centro de trabajo:",
        "afiliados)",
        "valor liquidado",
        "valor a pagar",
        "datos generales de la liquidacion",
        "datos generales de la liquidación",
        "entidad recaudo pagada",
        "ibc salud",
        "ibc pension",
        "ibc pensión",
        "valor pago",
        "estado planilla",
        "periodo salud",
        "periodo pensión",
        "periodo pension",
        "referencia de pago",
        "f. presentacion unica",
        "f. presentación única",
        "certificado de afiliacion positiva",
        "certificado de afiliación positiva",
    ]
    planilla_negative_markers = [
        "enviado el:",
        "datos adjuntos:",
        "asunto:",
        "formulario de afiliacion",
        "a. afiliacion",
        "a. afiliación",
        "b. traslado",
        "c. terminacion",
    ]
    planilla_hits = sum(1 for token in planilla_markers if token in haystack)
    if planilla_hits >= 2 and not any(token in haystack for token in planilla_negative_markers):
        return {"document_type": "soporte_ingresos", "legacy_code": 11, "code_source": "ocr_planilla_precise"}

    strong_rules: List[tuple[int, str, List[str], str]] = [
        (8, "rut", ["registro unico tributario", "r.u.t", " rut ", "direccion de impuestos y aduanas"], "ocr_rut"),
        (5, "camara_comercio", ["camara de comercio", "certificado de existencia", "matricula mercantil"], "ocr_camara"),
        (0, "formulario_afiliacion", ["formulario de afiliacion", "formulario unico de afiliacion", "cps-f-216"], "ocr_formulario"),
        (1, "anexo_sedes", ["sedes y centros de trabajo", "anexo formulario de afiliacion"], "ocr_sedes"),
        (2, "listado_trabajadores", ["listado de trabajadores", "trabajadores o estudiantes"], "ocr_listado"),
        (10, "entrega_documentos", ["comprobante entrega de documentos", "documentos anexos a la afiliacion"], "ocr_entrega"),
        (11, "soporte_pagos", ["recibo de pago", "ultimos recibos de pago", "pila pagada"], "ocr_pagos"),
        (12, "contrato", ["contrato de prestacion", "prestacion de servicios", "objeto del contrato"], "ocr_contrato"),
        (19, "identificacion_peligros", ["identificacion de peligros", "matriz de peligros"], "ocr_peligros"),
        (20, "examen_preocupacional", ["examen pre-ocupacional", "examen preocupacional"], "ocr_preocupacional"),
        (27, "beneficiario_final", ["beneficiario final"], "ocr_beneficiario"),
        (27, "beneficiario_final", ["participacion directa o indirecta mayor al 5", "participación directa o indirecta mayor al 5", "informacion de la compania", "información de la compañía"], "ocr_beneficiario_company"),
        (28, "sat", ["sistema de afiliacion transaccional", "canal sat"], "ocr_sat"),
        (98, "autorizacion", ["autorizacion clientes, proveedores y terceros", "autorizacion tratamiento de datos", "autorización de tratamiento de datos", "tratamiento de datos personales"], "ocr_autorizacion"),
    ]
    for code, doc_type, keys, label in strong_rules:
        if any(key in haystack for key in keys):
            return {"document_type": doc_type, "legacy_code": code, "code_source": label}

    if any(token in haystack for token in ["declaracion de renta", "declaracion renta", "honorarios", "ingresos", "desprendible de pago"]):
        return {"document_type": "soporte_ingresos", "legacy_code": 11, "code_source": "ocr_ingresos"}

    name_rules: List[tuple[int, str, List[str], str]] = [
        (6, "cedula", ["cedula", "cc_"], "name_cedula"),
        (5, "camara_comercio", ["camara", "comercio"], "name_camara"),
        (8, "rut", ["rut"], "name_rut"),
        (1, "anexo_sedes", ["sedes", "anexo_sedes"], "name_sedes"),
        (2, "listado_trabajadores", ["trabajadores", "listado"], "name_listado"),
        (10, "entrega_documentos", ["entrega", "anexos"], "name_entrega"),
        (11, "soporte_pagos", ["pagos", "recibo"], "name_pagos"),
        (12, "contrato", ["contrato"], "name_contrato"),
        (98, "autorizacion", ["autorizacion"], "name_autorizacion"),
    ]
    for code, doc_type, keys, label in name_rules:
        if any(key in name_txt for key in keys):
            return {"document_type": doc_type, "legacy_code": code, "code_source": label}

    if lower_name.endswith((".xlsx", ".xlsm", ".xls")):
        return {"document_type": "xlsx", "legacy_code": -1, "code_source": "file_xlsx"}
    if lower_name.endswith(".pdf"):
        return {"document_type": "pdf", "legacy_code": 99, "code_source": "file_pdf"}
    if lower_name.endswith((".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp")):
        return {"document_type": "imagen", "legacy_code": 99, "code_source": "file_image"}
    return {"document_type": "otro", "legacy_code": 99, "code_source": "fallback"}


def _infer_required_document_satisfaction(
    required_docs: List[str],
    docs: List[Dict[str, Any]],
    profile: Dict[str, Any],
) -> Dict[str, Dict[str, Any]]:
    xlsx_document = only_digits(profile.get("documento", ""))
    xlsx_nit = only_digits(profile.get("nit", ""))
    results: Dict[str, Dict[str, Any]] = {}
    grouped_types = {doc.get("document_type") for doc in docs}

    for required in required_docs:
        direct = required in grouped_types
        evidence: Dict[str, Any] = {"satisfied": direct, "direct": direct, "filename": "", "matched": "", "reason": ""}
        if direct:
            doc = next((item for item in docs if item.get("document_type") == required), None)
            evidence["filename"] = (doc or {}).get("filename", "")
            evidence["reason"] = "direct_type"
            results[required] = evidence
            continue

        if required == "cedula":
            for doc in docs:
                fields = doc.get("fields") or {}
                preview = normalize_haystack(doc.get("text_preview", ""))
                doc_type = str(doc.get("document_type") or "")
                candidates = [fields.get("representative_document", ""), fields.get("document_number", "")] + list(fields.get("all_numbers") or [])
                candidate = _best_numeric_candidate(xlsx_document, candidates) if xlsx_document else ""
                if candidate and (
                    doc_type not in {"carta", "rut", "camara_comercio", "soporte_ingresos", "entrega_documentos"}
                    and (
                        _looks_like_cedula_document(normalize_haystack(doc.get("filename", "")), preview)
                        or "numero de cedula" in preview
                        or "numero de identificacion cc" in preview
                        or "tipo de documento numero de identificacion cc" in preview
                        or "tipo de documento cc" in preview
                        or "identificacion personal" in preview
                        or ("lugar de nac" in preview and "expedicion" in preview)
                    )
                ):
                    evidence.update(
                        {
                            "satisfied": True,
                            "filename": doc.get("filename", ""),
                            "matched": candidate,
                            "reason": "inferred_identity_match",
                        }
                    )
                    break
        elif required == "rut":
            for doc in docs:
                fields = doc.get("fields") or {}
                preview = normalize_haystack(doc.get("text_preview", ""))
                nit_candidate = _best_numeric_candidate(xlsx_nit, [fields.get("nit", ""), fields.get("document_number", "")] + list(fields.get("all_numbers") or [])) if xlsx_nit else ""
                if nit_candidate and (
                    "dian" in preview
                    or "registro unico tributario" in preview
                    or "direccion de impuestos" in preview
                    or "numero de identificacion tributaria" in preview
                ):
                    evidence.update(
                        {
                            "satisfied": True,
                            "filename": doc.get("filename", ""),
                            "matched": nit_candidate,
                            "reason": "inferred_rut_match",
                        }
                    )
                    break
        elif required == "camara_comercio":
            for doc in docs:
                preview = normalize_haystack(doc.get("text_preview", ""))
                if _looks_like_camara_document(preview):
                    evidence.update(
                        {
                            "satisfied": True,
                            "filename": doc.get("filename", ""),
                            "reason": "inferred_camara_match",
                        }
                    )
                    break
        elif required == "soporte_ingresos":
            for doc in docs:
                fields = doc.get("fields") or {}
                preview = normalize_haystack(doc.get("text_preview", ""))
                if fields.get("has_income_hint") or any(
                    token in preview
                    for token in ["pila pagada", "recibo de pago", "desprendible de pago", "ingresos", "autorizacion cargue retroactivo"]
                ):
                    evidence.update(
                        {
                            "satisfied": True,
                            "filename": doc.get("filename", ""),
                            "reason": "inferred_income_match",
                        }
                    )
                    break

        results[required] = evidence
    return results


def _extract_fields(text: str) -> Dict[str, Any]:
    normalized = normalize_text(text)
    lowered = normalize_haystack(normalized)
    raw_number_groups = re.findall(r"\d[\d.,\-\s]{4,}\d", normalized)
    digits: List[str] = []
    for item in raw_number_groups:
        cleaned = only_digits(item)
        if 6 <= len(cleaned) <= 15 and cleaned not in digits:
            digits.append(cleaned)
    for item in re.findall(r"\b\d{6,15}\b", normalized):
        cleaned = only_digits(item)
        if 6 <= len(cleaned) <= 15 and cleaned not in digits:
            digits.append(cleaned)
    doc_number = digits[0] if digits else ""
    nit_match = re.search(r"\bnit[^0-9]{0,12}(\d{6,12})\b", lowered) or re.search(r"\b(\d{9,10})\b", normalized)
    date_matches = re.findall(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b", normalized)
    has_rut = "rut" in lowered or "nit" in lowered or "tributario" in lowered
    company_name = ""
    rep_name = ""
    rep_doc = ""
    match = re.search(r"raz[oó]n social[^A-Za-z0-9]{0,10}([A-ZÁÉÍÓÚÑ0-9 .,&-]{6,120})", normalized, flags=re.IGNORECASE)
    if match:
        candidate = normalize_text(match.group(1))
        if _looks_like_company_name(candidate):
            company_name = candidate
    if not company_name:
        match = re.search(r"(PALMAS\s+DE\s+PUERTO\s+GAITAN(?:\s+S\.\s*A\.\s*S\.?)?)", normalized, flags=re.IGNORECASE)
        if match:
            company_name = normalize_text(match.group(1))
    match = re.search(r"apellidos y nombres del representante legal[^A-Za-z0-9]{0,20}([A-ZÁÉÍÓÚÑ ]{6,120})", normalized, flags=re.IGNORECASE)
    if match:
        candidate = normalize_text(match.group(1))
        candidate = _normalize_person_name(candidate)
        if _looks_like_person_name(candidate):
            rep_name = candidate
    if not rep_name:
        match = re.search(r"\byo\s+([A-ZÁÉÍÓÚÑ ]{6,120})\s+identificad[oa]?\s+con", normalized, flags=re.IGNORECASE)
        if match:
            candidate = normalize_text(match.group(1))
            candidate = _normalize_person_name(candidate)
            if _looks_like_person_name(candidate):
                rep_name = candidate
    match = re.search(r"n[uú]mero de documento[^0-9]{0,20}([\d.,\-\s]{6,20})", normalized, flags=re.IGNORECASE)
    if match:
        rep_doc = only_digits(match.group(1))
    if (
        "cedula de ciudadania" in lowered
        or "cedula de" in lowered
        or "cédula de" in normalized.lower()
        or "identificacion personal" in lowered
        or "nuip" in lowered
    ):
        cedula_number_match = re.search(
            r"(?:nuip|c[eé]dula\s+de\s+ciudadan[ií]a)[^\d]{0,12}(\d{1,3}(?:[.\s]\d{3}){1,3})",
            normalized,
            flags=re.IGNORECASE,
        )
        if not cedula_number_match:
            cedula_number_match = re.search(
                r"\b(\d{1,3}(?:[.\s]\d{3}){1,3})\s+(?:ENE|FEB|MAR|ABR|MAY|JUN|JUL|AGO|SEP|SEPT|OCT|NOV|DIC)\b",
                normalized,
                flags=re.IGNORECASE,
            )
        if cedula_number_match:
            rep_doc = only_digits(cedula_number_match.group(1))
        cedula_name = _extract_name_from_cedula_text(normalized)
        cedula_name = _normalize_person_name(cedula_name)
        if _looks_like_person_name(cedula_name):
            rep_name = cedula_name
        cedula_numbers = [only_digits(item) for item in digits if 6 <= len(only_digits(item)) <= 15]
        if cedula_numbers:
            rep_doc = rep_doc or cedula_numbers[0]
            doc_number = rep_doc or doc_number
    textual_issue_date = _parse_spanish_date_text(normalized)
    digit_candidates = [only_digits(item) for item in digits]
    filtered_nit_candidates = [
        item for item in digit_candidates
        if 8 <= len(item) <= 10 and not item.startswith("00")
    ]
    nit_value = nit_match.group(1) if nit_match and has_rut else ""
    if nit_value and len(only_digits(nit_value)) > 10:
        nit_value = ""
    if not nit_value and filtered_nit_candidates:
        nit_value = filtered_nit_candidates[0]

    return {
        "document_number": doc_number,
        "nit": nit_value,
        "all_numbers": digits[:10],
        "dates": date_matches[:5],
        "company_name": company_name,
        "representative_name": rep_name,
        "representative_document": rep_doc,
        "issue_date_textual": textual_issue_date.strftime("%Y-%m-%d") if textual_issue_date else "",
        "has_signature_hint": "firma" in lowered,
        "has_income_hint": any(token in lowered for token in ["ingres", "salario", "honorarios", "renta"]),
        "has_contrato_hint": "contrato" in lowered or "prestacion de servicios" in lowered,
        "has_worker_list_hint": "trabajador" in lowered or "trabajadores" in lowered,
    }


def _parse_date(text: str) -> Optional[datetime]:
    match = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})\b", text)
    if not match:
        return None
    day, month, year = match.groups()
    year_num = int(year)
    if year_num < 100:
        year_num += 2000
    try:
        return datetime(year_num, int(month), int(day))
    except ValueError:
        return None


def _parse_spanish_date_text(text: str) -> Optional[datetime]:
    lowered = normalize_haystack(text)
    month_map = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "setiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }
    explicit_patterns = [
        r"fecha\s+de\s+expedicion[:\s]+(\d{1,2})[/-](\d{1,2})[/-](\d{4})",
        r"fecha\s+expedicion[:\s]+(\d{1,2})[/-](\d{1,2})[/-](\d{4})",
        r"fecha\s+de\s+expedicion[:\s]+(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})",
        r"fecha\s+expedicion[:\s]+(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})",
    ]
    for pattern in explicit_patterns:
        match = re.search(pattern, lowered)
        if not match:
            continue
        groups = match.groups()
        try:
            if len(groups) == 3 and groups[1].isdigit():
                day, month, year = groups
                return datetime(int(year), int(month), int(day))
            day, month_name, year = groups
            month = month_map.get(month_name)
            if month:
                return datetime(int(year), int(month), int(day))
        except ValueError:
            pass

    match = re.search(r"(\d{1,2})\s+de\s+([a-z]+)\s+de\s+(\d{4})", lowered)
    if not match:
        return None
    day, month_name, year = match.groups()
    month = month_map.get(month_name)
    if not month:
        return None
    try:
        return datetime(int(year), int(month), int(day))
    except ValueError:
        return None


def _ocr_image(path: Path) -> Dict[str, Any]:
    with Image.open(path) as image:
        text = _ocr_best_text_from_image(image)
    return {
        "text": normalize_text(text),
        "used_ocr": True,
        "pages_processed": 1,
    }


def _prepare_ocr_variants(image: Image.Image) -> List[Image.Image]:
    base = image.convert("L")
    enlarged = base.resize((max(base.width * 2, 1), max(base.height * 2, 1)), Image.Resampling.LANCZOS)
    sharpened = enlarged.filter(ImageFilter.SHARPEN)
    contrasted = ImageEnhance.Contrast(sharpened).enhance(2.6)
    autocontrasted = ImageOps.autocontrast(contrasted)
    thresholded = autocontrasted.point(lambda px: 255 if px > 165 else 0, mode="1").convert("L")
    soft_thresholded = autocontrasted.point(lambda px: 255 if px > 145 else 0, mode="1").convert("L")
    return [enlarged, autocontrasted, thresholded, soft_thresholded]


def _ocr_best_text_from_image(image: Image.Image) -> str:
    variants = _prepare_ocr_variants(image)
    configs = [
        "--psm 6",
        "--psm 11",
    ]
    best_text = ""
    best_score = -1
    for variant in variants:
        for config in configs:
            try:
                text = normalize_text(pytesseract.image_to_string(variant, lang="spa+eng", config=config))
            except Exception:
                continue
            if not text:
                continue
            alpha = sum(1 for ch in text if ch.isalpha())
            digits = sum(1 for ch in text if ch.isdigit())
            # Favor useful OCR, with extra weight for identity-like documents that contain many digits.
            score = len(text) + alpha + (digits * 2)
            if "cedula" in normalize_haystack(text) or "identificacion" in normalize_haystack(text):
                score += 40
            if score > best_score:
                best_score = score
                best_text = text
    return best_text


def _text_quality_is_low(text: str) -> bool:
    cleaned = normalize_text(text)
    if len(cleaned) < 80:
        return True
    alpha = sum(1 for ch in cleaned if ch.isalpha())
    digits = sum(1 for ch in cleaned if ch.isdigit())
    spaces = sum(1 for ch in cleaned if ch.isspace())
    useful = alpha + digits + spaces
    if useful == 0:
        return True
    alpha_ratio = alpha / useful
    digit_ratio = digits / useful
    if alpha_ratio < 0.22:
        return True
    if digit_ratio > 0.55 and alpha_ratio < 0.35:
        return True
    if cleaned.count("/") > 20:
        return True
    return False


def format_reason_lines(message: str) -> List[str]:
    text = normalize_text(message)
    if not text:
        return []
    if "El XLSX trae trabajadores duplicados por documento:" in text:
        prefix, details = text.split(":", 1)
        entries = [item.strip(" .") for item in re.split(r"\s*;\s*", details) if item.strip(" .")]
        lines = [f"{prefix}:"]
        lines.extend(f"  - {item}" for item in entries)
        return lines
    if "Detalle por hoja:" in text:
        prefix, details = text.split("Detalle por hoja:", 1)
        entries = [item.strip(" .") for item in re.split(r"\s*,\s*", details) if item.strip(" .")]
        lines = [normalize_text(prefix).rstrip(" .")]
        lines.extend(f"  - {item}" for item in entries)
        return lines
    return [text]


def _read_pdf(path: Path) -> Dict[str, Any]:
    text_parts: List[str] = []
    extracted_pages = 0
    try:
        reader = PdfReader(str(path))
        for page in reader.pages[:5]:
            extracted = normalize_text(page.extract_text() or "")
            if extracted:
                text_parts.append(extracted)
            extracted_pages += 1
    except Exception:
        text_parts = []

    joined_text = "\n".join(text_parts)
    if text_parts and not _text_quality_is_low(joined_text):
        return {
            "text": joined_text,
            "used_ocr": False,
            "pages_processed": max(min(extracted_pages, 5), len(text_parts)),
        }

    images = convert_from_path(str(path), first_page=1, last_page=3, dpi=220)
    ocr_parts = [normalize_text(_ocr_best_text_from_image(image)) for image in images]
    return {
        "text": "\n".join(part for part in ocr_parts if part),
        "used_ocr": True,
        "pages_processed": len(images),
    }


def _explode_multipage_pdf_bytes(filename: str, content: bytes, max_pages: int = 300) -> List[tuple[str, bytes]]:
    safe_name = Path(str(filename or "")).name
    if not safe_name.lower().endswith(".pdf"):
        return [(safe_name, content)]
    try:
        reader = PdfReader(io.BytesIO(content))
    except Exception:
        return [(safe_name, content)]
    total_pages = len(reader.pages)
    if total_pages <= 1:
        return [(safe_name, content)]

    stem = Path(safe_name).stem
    exploded: List[tuple[str, bytes]] = []
    for index, page in enumerate(reader.pages[: max(1, min(total_pages, max_pages))], start=1):
        writer = PdfWriter()
        writer.add_page(page)
        buffer = io.BytesIO()
        writer.write(buffer)
        exploded.append((f"{stem}__p{index:03d}.pdf", buffer.getvalue()))
    return exploded


def _extract_worker_records_from_rows(rows: List[tuple[Any, ...]]) -> tuple[List[Dict[str, str]], int]:
    def _header_key(value: Any) -> str:
        return normalize_haystack(normalize_text(value)).replace(" ", "_")

    header_index = -1
    headers: List[str] = []
    best_score = -1
    for idx, row in enumerate(rows[:50]):
        values = [normalize_text(cell) for cell in row[:40]]
        normalized = [_header_key(cell) for cell in values if cell]
        has_document = any(
            key in normalized
            for key in ["documento", "numero_documento", "num_id_trabajador", "numero_de_identificacion"]
        )
        has_name = any(
            key in normalized
            for key in ["nombre", "nombre_trabajador", "primer_nombre", "primer_apellido", "segundo_apellido"]
        )
        worker_markers = sum(
            1
            for key in [
                "cargo",
                "salario",
                "eps",
                "pension",
                "correo_electronico",
                "municipio_distrito",
                "tipo_de_salario",
                "tipo_de_trabajador",
            ]
            if key in normalized
        )
        if has_document and has_name:
            score = worker_markers
            if "numero_de_identificacion" in normalized:
                score += 2
            if "tipo_de_documento" in normalized:
                score += 1
            if score > best_score:
                best_score = score
                header_index = idx
                headers = [_header_key(cell) for cell in values]

    records: List[Dict[str, str]] = []
    if header_index >= 0 and headers:
        for row_offset, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            values = [normalize_text(cell) for cell in row[: len(headers)]]
            if not any(values):
                continue
            record = {headers[pos]: values[pos] for pos in range(min(len(headers), len(values))) if headers[pos]}
            document_value = (
                record.get("documento")
                or record.get("numero_documento")
                or record.get("num_id_trabajador")
                or record.get("numero_de_identificacion")
                or ""
            )
            name_value = (
                record.get("nombre")
                or record.get("nombre_trabajador")
                or record.get("primer_nombre")
                or record.get("primer_apellido")
                or ""
            )
            doc_digits = only_digits(document_value)
            doc_type = normalize_haystack(record.get("tipo_de_documento", "")).upper()
            if normalize_haystack(document_value) in {"total", "total_centros_de_trabajo"}:
                break
            if document_value or name_value:
                if doc_digits and len(doc_digits) >= 5 and doc_type in {"CC", "CE", "CD", "SC", "PE", "PT", "RC", "TI", "NI"}:
                    record["numero_de_identificacion"] = doc_digits
                    record["_row"] = str(row_offset)
                    records.append(record)
                elif not document_value and name_value:
                    record["_row"] = str(row_offset)
                    records.append(record)
    return records, header_index


def _read_xlsx(path: Path) -> Dict[str, Any]:
    workbook = load_workbook(path, data_only=True)
    sheets: List[Dict[str, Any]] = []
    flat_pairs: Dict[str, str] = {}
    records: List[Dict[str, str]] = []
    worker_sheet_counts: Dict[str, int] = {}
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        preview = [[normalize_text(cell) for cell in row[:12]] for row in rows[:8]]
        sheets.append({"name": sheet.title, "preview": preview})
        for row in [[normalize_text(cell) for cell in raw_row[:12]] for raw_row in rows[:80]]:
            if len(row) >= 2 and row[0] and row[1]:
                key = normalize_haystack(normalize_text(row[0])).replace(" ", "_")
                if key not in flat_pairs:
                    flat_pairs[key] = row[1]
        sheet_records, header_index = _extract_worker_records_from_rows(rows)
        if header_index >= 0:
            worker_sheet_counts[sheet.title] = len(
                [record for record in sheet_records if only_digits(record.get("numero_de_identificacion", ""))]
            )
        for record in sheet_records:
            if sheet.title:
                record["_sheet"] = sheet.title
            records.append(record)
    profile = {
        "tipo_afiliado": flat_pairs.get("tipo_afiliado") or flat_pairs.get("tipoafiliacion") or "",
        "documento": flat_pairs.get("documento") or flat_pairs.get("numero_documento") or "",
        "nombre": flat_pairs.get("nombre") or flat_pairs.get("nombre_trabajador") or "",
        "empresa": flat_pairs.get("empresa") or flat_pairs.get("empleador") or flat_pairs.get("razon_social") or "",
        "nit": flat_pairs.get("nit") or flat_pairs.get("nit_empleador") or "",
        "documento_empleador": flat_pairs.get("numerodocumentoempleador") or flat_pairs.get("documento_empleador") or flat_pairs.get("id_empresa") or "",
        "lote": flat_pairs.get("lote") or flat_pairs.get("nl") or flat_pairs.get("numero_lote") or "",
        "idtramite": flat_pairs.get("idtramite") or flat_pairs.get("id_tramite") or "",
    }

    clean_preview = _generate_clean_from_workbook(workbook, path.name)
    contract_fields = _extract_employer_from_contract_text((clean_preview.get("contrato_clean") or {}).get("content", ""))
    company_name = normalize_text(contract_fields.get("empresa", ""))
    nit_value = only_digits(contract_fields.get("nit", ""))
    rep_name = normalize_text(contract_fields.get("representante_legal", ""))
    rep_doc = only_digits(contract_fields.get("doc_representante", ""))
    if company_name and len(company_name) >= 5:
        profile["empresa"] = profile["empresa"] or company_name
    if 8 <= len(nit_value) <= 12:
        profile["nit"] = profile["nit"] or nit_value
        profile["documento_empleador"] = profile["documento_empleador"] or nit_value
    if rep_name and len(rep_name) >= 5:
        profile["nombre"] = profile["nombre"] or rep_name
    if 6 <= len(rep_doc) <= 15:
        profile["documento"] = profile["documento"] or rep_doc

    return {
        "sheets": sheets,
        "profile": profile,
        "flat_pairs": flat_pairs,
        "records": records,
        "worker_sheet_counts": worker_sheet_counts,
    }


def _cell_to_clean_text(value: Any, force_int_float: bool = False) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%dT%H:%M:%S")
    if isinstance(value, float):
        if abs(value - int(value)) < 1e-9:
            return str(int(value))
        if force_int_float:
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return text if text else "NULL"


def _sheet_to_clean_lines(sheet: Any, force_int_float: bool = False) -> List[str]:
    lines: List[str] = []
    for row in sheet.iter_rows(values_only=True):
        values = [_cell_to_clean_text(value, force_int_float=force_int_float) for value in row]
        first = -1
        last = -1
        for index, item in enumerate(values):
            if item != "NULL":
                if first < 0:
                    first = index
                last = index
        if first >= 0 and last >= 0:
            lines.append("|".join(values[first : last + 1]))
    return lines


def _generate_clean_from_workbook(workbook: Any, excel_filename: str) -> Dict[str, Any]:
    def _norm(value: str) -> str:
        return re.sub(r"\s+", " ", str(value or "").strip().lower())

    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        return {"ok": False, "message": "Excel sin hojas."}

    main_sheet = ""
    for name in sheet_names:
        norm_name = _norm(name)
        if "formulario" in norm_name and "afili" in norm_name:
            main_sheet = name
            break
    if not main_sheet:
        main_sheet = sheet_names[0]

    worker_sheets: List[str] = []
    indep_sheet = ""
    for name in sheet_names:
        norm_name = _norm(name)
        if "sede" in norm_name and "trabajador" in norm_name:
            worker_sheets.append(name)
        if ("independiente" in norm_name and "723" in norm_name) or norm_name == "independientes 723":
            indep_sheet = name

    contrato_lines = _sheet_to_clean_lines(workbook[main_sheet], force_int_float=False)
    workers_multi: List[Dict[str, Any]] = []
    for idx, name in enumerate(worker_sheets, start=1):
        sheet = workbook[name]
        rows = list(sheet.iter_rows(values_only=True))
        sheet_records, _ = _extract_worker_records_from_rows(rows)
        worker_rows = [record for record in sheet_records if only_digits(record.get("numero_de_identificacion", ""))]
        if not worker_rows:
            continue
        sheet_lines = _sheet_to_clean_lines(sheet, force_int_float=True)
        if not sheet_lines:
            continue
        sede_match = re.search(r"sede\s*0*(\d+)", name, flags=re.IGNORECASE)
        sede_num = int(sede_match.group(1)) if sede_match else idx
        workers_multi.append(
            {
                "sheet": name,
                "filename": f"Sede{sede_num:02d}-Trabajadores_clean.txt",
                "lines": len(sheet_lines),
                "workers": len(worker_rows),
                "content": "\n".join(sheet_lines) + "\n",
                "preview": sheet_lines[:12],
            }
        )

    independientes_block = None
    if indep_sheet:
        indep_lines = _sheet_to_clean_lines(workbook[indep_sheet], force_int_float=True)
        if indep_lines:
            independientes_block = {
                "sheet": indep_sheet,
                "filename": "independientes_clean.txt",
                "lines": len(indep_lines),
                "content": "\n".join(indep_lines) + "\n",
                "preview": indep_lines[:12],
            }

    safe_name = re.sub(r"[^0-9A-Za-z_-]+", "", Path(str(excel_filename or "contrato")).stem) or "contrato"
    for idx, item in enumerate(workers_multi, start=1):
        sede_match = re.search(r"sede\s*0*(\d+)", str(item.get("sheet") or ""), flags=re.IGNORECASE)
        sede_num = int(sede_match.group(1)) if sede_match else idx
        item["filename"] = f"Sede{sede_num:02d}-Trabajadores_{safe_name}_clean.txt"

    return {
        "ok": bool(contrato_lines),
        "source_sheets": {"main_sheet": main_sheet, "worker_sheets": worker_sheets},
        "contrato_clean": {
            "filename": f"contrato_{safe_name}_clean.txt",
            "lines": len(contrato_lines),
            "content": "\n".join(contrato_lines) + ("\n" if contrato_lines else ""),
            "preview": contrato_lines[:12],
        },
        "trabajadores_clean_multi": workers_multi,
        "independientes_clean": independientes_block,
    }


def _extract_employer_from_contract_text(contract_text: str) -> Dict[str, str]:
    text = str(contract_text or "")
    if not text.strip():
        return {}
    parts = [part.strip() for part in text.split("|")]
    norm_parts = [normalize_haystack(part) for part in parts]

    def by_label(*labels: str) -> str:
        aliases = [normalize_haystack(label) for label in labels if label]
        for strict in (True, False):
            for index, label in enumerate(norm_parts[:-1]):
                matched = any(label == alias or label.startswith(alias) for alias in aliases) if strict else any(alias in label for alias in aliases)
                if matched:
                    for probe in range(index + 1, min(index + 20, len(parts))):
                        value = str(parts[probe] or "").strip()
                        norm_value = normalize_haystack(value)
                        if not value or norm_value in {"null", "na", "n a"}:
                            continue
                        if "tipo de documento" in norm_value and len(value) <= 3:
                            continue
                        return value
        return ""

    rep_name = ""
    for index, label in enumerate(norm_parts):
        if "apellidos y nombres del representante legal" in label:
            values: List[str] = []
            for probe in range(index + 1, min(index + 8, len(parts))):
                value = str(parts[probe] or "").strip()
                norm_value = normalize_haystack(value)
                if not value or norm_value in {"null", "na", "n a"}:
                    continue
                if any(token in norm_value for token in ["tipo de documento", "numero de documento", "correo electronico"]):
                    break
                values.append(value)
            rep_name = " ".join(values).strip()
            break

    return {
        "empresa": by_label("1. Apellidos y nombres o razón social"),
        "nit": only_digits(by_label("3. Número de documento o NIT", "nit")),
        "representante_legal": rep_name,
        "doc_representante": only_digits(by_label("6. Número de documento")),
        "direccion_empleador": by_label("Dirección de la sede principal", "direccion"),
        "telefono_empleador": only_digits(by_label("Teléfono fijo/celular")),
        "ciudad_empleador": only_digits(by_label("Municipio/Distrito", "ciudad")),
        "zona_empleador": by_label("Zona"),
        "correo_empleador": by_label("Correo electrónico"),
        "numero_sedes": only_digits(by_label("Número de sedes")),
        "numero_trabajadores": only_digits(by_label("Número total de trabajadores o estudiantes", "Cantidad de trabajadores y estudiantes")),
    }


def _legacy_nova_url() -> str:
    configured = str(settings.legacy_backend_url or "").strip().rstrip("/")
    if not configured:
        return ""
    return configured[:-13] + "/nova" if configured.endswith("/afiliaciones") else configured


def _generate_clean_via_legacy_nova(excel_filename: str, excel_bytes: bytes) -> Dict[str, Any]:
    backend_url = _legacy_nova_url()
    if not backend_url:
        return {"ok": False, "message": "legacy_backend_url no configurado."}
    if not excel_bytes:
        return {"ok": False, "message": "El Excel está vacío."}

    candidate_urls = [backend_url]
    if "127.0.0.1" in backend_url:
        candidate_urls.append(backend_url.replace("127.0.0.1", "host.docker.internal"))
    if "localhost" in backend_url:
        candidate_urls.append(backend_url.replace("localhost", "host.docker.internal"))

    last_error = ""
    for base_url in candidate_urls:
        try:
            files = {
                "excel_file": (
                    excel_filename,
                    excel_bytes,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            }
            response = httpx.post(f"{base_url}/precheck/generate-clean-upload", files=files, timeout=180.0)
            if response.status_code == 200:
                out = response.json()
                if isinstance(out, dict):
                    out["_source"] = f"legacy_nova:{base_url}"
                return out
            last_error = f"HTTP {response.status_code}: {response.text[:600]}"
        except Exception as exc:
            last_error = f"{type(exc).__name__}: {exc}"
    return {"ok": False, "message": last_error or "No pude generar clean con el clone NOVA."}


def _infer_tipo_afiliado_from_docs(docs: List[Dict[str, Any]]) -> str:
    haystack = normalize_haystack(" ".join(str(item.get("text_preview") or "") for item in docs[:10]))
    if "traslado" in haystack:
        return "traslado"
    if "nueva" in haystack or "afiliacion" in haystack:
        return "nueva"
    return ""


def _enrich_xlsx_profile_from_clean(
    xlsx_profile: Dict[str, Any],
    clean_output: Dict[str, Any],
    docs: List[Dict[str, Any]],
) -> Dict[str, Any]:
    profile = dict((xlsx_profile or {}).get("profile") or {})
    flat_pairs = dict((xlsx_profile or {}).get("flat_pairs") or {})
    worker_sheet_counts = dict((xlsx_profile or {}).get("worker_sheet_counts") or {})
    seeded_employer_nit = _normalize_company_nit(profile.get("documento_empleador") or profile.get("nit", ""), docs)
    if 8 <= len(seeded_employer_nit) <= 12:
        profile["documento_empleador"] = seeded_employer_nit
        profile["nit"] = seeded_employer_nit
    contrato_clean = clean_output.get("contrato_clean") if isinstance(clean_output.get("contrato_clean"), dict) else {}
    contract_fields = _extract_employer_from_contract_text(str(contrato_clean.get("content") or ""))

    company_name = normalize_text(contract_fields.get("empresa", ""))
    nit_value = only_digits(contract_fields.get("nit", ""))
    rep_name = normalize_text(contract_fields.get("representante_legal", ""))
    rep_doc = only_digits(contract_fields.get("doc_representante", ""))
    tipo_afiliado = normalize_text(profile.get("tipo_afiliado", ""))

    if not company_name:
        for doc in docs:
            doc_nit = only_digits((doc.get("fields") or {}).get("nit", ""))
            preview = normalize_text(doc.get("text_preview", ""))
            if doc_nit and len(doc_nit) >= 8 and "palmas de puerto gaitan" in normalize_haystack(preview):
                company_name = "PALMAS DE PUERTO GAITAN S.A.S"
                break
    if not nit_value:
        for doc in docs:
            doc_nit = only_digits((doc.get("fields") or {}).get("nit", ""))
            if 8 <= len(doc_nit) <= 12:
                nit_value = doc_nit
                break
    if not nit_value:
        for doc in docs:
            if doc.get("document_type") not in {"rut", "camara_comercio", "formulario_afiliacion", "cedula"}:
                continue
            for raw in ((doc.get("fields") or {}).get("all_numbers") or []):
                digits = only_digits(raw)
                if 8 <= len(digits) <= 12:
                    nit_value = digits
                    break
            if nit_value:
                break
    if not rep_doc:
        for doc in docs:
            doc_num = only_digits((doc.get("fields") or {}).get("document_number", ""))
            if 6 <= len(doc_num) <= 15 and doc.get("document_type") in {"cedula", "formulario_afiliacion"}:
                rep_doc = doc_num
                break
    if not rep_name:
        for doc in docs:
            preview = normalize_text(doc.get("text_preview", ""))
            match = re.search(r"([A-ZÁÉÍÓÚÑ]+(?:\s+[A-ZÁÉÍÓÚÑ]+){2,5})", preview, flags=re.IGNORECASE)
            if match:
                rep_name = normalize_text(match.group(0))
                break
    if not tipo_afiliado:
        tipo_afiliado = _infer_tipo_afiliado_from_docs(docs)

    if company_name and len(company_name) >= 5:
        profile["empresa"] = company_name
    employer_document_hint = _normalize_company_nit(profile.get("documento_empleador", ""), docs)
    if 8 <= len(employer_document_hint) <= 12:
        profile["nit"] = employer_document_hint
        profile["documento_empleador"] = employer_document_hint
    elif 8 <= len(nit_value) <= 12:
        profile["nit"] = nit_value
        profile["documento_empleador"] = profile.get("documento_empleador") or nit_value
    if rep_name and len(rep_name) >= 5:
        profile["nombre"] = _normalize_person_name(rep_name)
    if 6 <= len(rep_doc) <= 15:
        profile["documento"] = rep_doc
    if tipo_afiliado:
        profile["tipo_afiliado"] = tipo_afiliado
    active_worker_sedes = sum(1 for count in worker_sheet_counts.values() if int(count or 0) > 0)
    total_worker_rows = sum(int(count or 0) for count in worker_sheet_counts.values())
    explicit_worker_total = only_digits(
        flat_pairs.get("numero_total_de_trabajadores_o_estudiantes")
        or flat_pairs.get("cantidad_de_trabajadores_y_estudiantes")
        or ""
    )
    explicit_sedes_total = only_digits(flat_pairs.get("numero_de_sedes") or "")
    clean_sede_files = len(clean_output.get("trabajadores_clean_multi") or [])
    if explicit_sedes_total:
        profile["numero_sedes"] = int(explicit_sedes_total)
    elif active_worker_sedes > 0:
        profile["numero_sedes"] = active_worker_sedes
    elif clean_sede_files > 0:
        profile["numero_sedes"] = clean_sede_files
    elif only_digits(contract_fields.get("numero_sedes", "")):
        profile["numero_sedes"] = int(only_digits(contract_fields.get("numero_sedes", "")))
    if explicit_worker_total:
        profile["numero_trabajadores"] = int(explicit_worker_total)
    elif total_worker_rows > 0:
        profile["numero_trabajadores"] = total_worker_rows
    elif only_digits(contract_fields.get("numero_trabajadores", "")):
        profile["numero_trabajadores"] = int(only_digits(contract_fields.get("numero_trabajadores", "")))

    enriched = dict(xlsx_profile or {})
    enriched["profile"] = profile
    if clean_output:
        enriched["clean_source"] = clean_output.get("_source", "local")
        enriched["clean_preview"] = {
            "contrato_lines": int((contrato_clean or {}).get("lines") or 0),
            "sede_files": len(clean_output.get("trabajadores_clean_multi") or []),
            "has_independientes": bool((clean_output.get("independientes_clean") or {}).get("content")),
        }
    return enriched


def _finalize_profile_from_docs(xlsx_profile: Dict[str, Any], docs: List[Dict[str, Any]]) -> Dict[str, Any]:
    profile = dict((xlsx_profile or {}).get("profile") or {})
    cedula_doc = next((doc for doc in docs if doc.get("document_type") == "cedula"), None)
    employer_document_hint = _normalize_company_nit(profile.get("documento_empleador", ""), docs)

    if not _looks_like_company_name(profile.get("empresa", "")):
        profile["empresa"] = ""
    if not profile.get("empresa"):
        for preferred in ["camara_comercio", "formulario_afiliacion", "rut", "cedula"]:
            for doc in docs:
                if doc.get("document_type") != preferred:
                    continue
                candidate = normalize_text((doc.get("fields") or {}).get("company_name", ""))
                if candidate:
                    candidate = re.split(r"\b(?:nit|cc|numero|número)\b", candidate, maxsplit=1, flags=re.IGNORECASE)[0].strip(" .,-")
                if _looks_like_company_name(candidate):
                    profile["empresa"] = candidate
                    break
            if profile.get("empresa"):
                break

    nit_votes: Dict[str, int] = {}
    weighted_types = {
        "rut": 5,
        "camara_comercio": 4,
        "formulario_afiliacion": 3,
        "entrega_documentos": 3,
        "soporte_ingresos": 2,
        "constancia_afiliacion": 2,
    }
    for doc in docs:
        fields = doc.get("fields") or {}
        preview = normalize_haystack(doc.get("text_preview", ""))
        doc_type = doc.get("document_type") or ""
        if any(
            token in preview
            for token in [
                "matriz o controlante",
                "nit de la matriz o controlante",
                "sociedades y organismos extranjeros",
                "revisor fiscal",
                "contador",
                "establecimientos, agencias, sucursales",
                "fecha inicio ejercicio representacion",
                "fecha inicio ejercicio representación",
            ]
        ):
            continue
        candidates = [fields.get("nit", "")] + list(fields.get("all_numbers") or [])
        for raw in candidates:
            digits = only_digits(raw)
            if not (8 <= len(digits) <= 10):
                continue
            if digits.startswith("00"):
                continue
            weight = weighted_types.get(doc_type, 1)
            if digits == only_digits(fields.get("document_number", "")) and doc_type in {"cedula", "anexo_sedes"}:
                continue
            if any(token in preview for token in ["nit", "identificacion", "identificación", "razon social", "razón social", "pagada ni"]):
                weight += 1
            nit_votes[digits] = nit_votes.get(digits, 0) + weight
    if 8 <= len(employer_document_hint) <= 12:
        profile["nit"] = employer_document_hint
    elif nit_votes:
        best_nit = sorted(nit_votes.items(), key=lambda item: (item[1], len(item[0])), reverse=True)[0][0]
        current_nit = _normalize_company_nit(profile.get("nit", ""), docs)
        if not current_nit or current_nit not in nit_votes or nit_votes.get(best_nit, 0) > nit_votes.get(current_nit, 0):
            profile["nit"] = _normalize_company_nit(best_nit, docs)
    normalized_profile_nit = _normalize_company_nit(profile.get("nit", ""), docs)
    if 8 <= len(normalized_profile_nit) <= 12:
        profile["nit"] = normalized_profile_nit
        profile["documento_empleador"] = employer_document_hint or normalized_profile_nit

    current_name = normalize_text(profile.get("nombre", ""))
    if not _looks_like_person_name(current_name):
        for preferred in ["cedula", "carta", "constancia_afiliacion", "formulario_afiliacion", "anexo_sedes"]:
            for doc in docs:
                if doc.get("document_type") != preferred:
                    continue
                fields = doc.get("fields") or {}
                candidate = normalize_text(fields.get("representative_name", ""))
                candidate = _normalize_person_name(candidate)
                if _looks_like_person_name(candidate) and "cps" not in normalize_haystack(candidate):
                    profile["nombre"] = candidate
                    break
            if _looks_like_person_name(profile.get("nombre", "")):
                break
    if cedula_doc:
        cedula_name = normalize_text((cedula_doc.get("fields") or {}).get("representative_name", ""))
        cedula_name = _normalize_person_name(cedula_name)
        if _looks_like_person_name(cedula_name):
            profile["nombre"] = cedula_name

    formulario_rep_doc = ""
    for doc in docs:
        if doc.get("document_type") != "formulario_afiliacion":
            continue
        formulario_rep_doc = only_digits((doc.get("fields") or {}).get("representative_document", "") or (doc.get("fields") or {}).get("document_number", ""))
        if formulario_rep_doc:
            break
    current_doc = only_digits(profile.get("documento", ""))
    if 6 <= len(formulario_rep_doc) <= 15 and (len(current_doc) < 8 or current_doc == "0500100"):
        profile["documento"] = formulario_rep_doc

    enriched = dict(xlsx_profile or {})
    enriched["profile"] = profile
    return enriched


def _build_required_documents(xlsx_profile: Dict[str, Any]) -> List[str]:
    afiliado = normalize_text(xlsx_profile.get("tipo_afiliado", "")).lower()
    required = ["cedula", "rut", "soporte_ingresos"]
    if any(token in afiliado for token in ["independ", "contratista"]):
        required.append("contrato")
    return required


def _doc_by_type(docs: List[Dict[str, Any]], document_type: str) -> List[Dict[str, Any]]:
    return [item for item in docs if item.get("document_type") == document_type]


def _unique_preserve(values: List[str]) -> List[str]:
    seen = set()
    output: List[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            output.append(value)
    return output


def _parse_nomina_value(value: Any) -> int:
    raw = normalize_text(value)
    if not raw:
        return 0
    digits = re.sub(r"[^\d]", "", raw)
    return int(digits) if digits else 0


def _summarize_received_documents(docs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summary: Dict[str, Dict[str, Any]] = {}
    for doc in docs:
        key = doc.get("document_type", "otro")
        bucket = summary.setdefault(
            key,
            {
                "document_type": key,
                "label": DOC_TYPE_LABELS.get(key, key.replace("_", " ").title()),
                "count": 0,
                "legacy_codes": set(),
                "files": [],
            },
        )
        bucket["count"] += 1
        bucket["files"].append(doc.get("filename"))
        code = doc.get("legacy_code")
        if isinstance(code, int):
            bucket["legacy_codes"].add(code)
    rows = []
    for bucket in summary.values():
        bucket["legacy_codes"] = sorted(bucket["legacy_codes"])
        rows.append(bucket)
    rows.sort(key=lambda item: item["document_type"])
    return rows


def _build_precheck_summary(xlsx_profile: Dict[str, Any], docs: List[Dict[str, Any]], required_docs: List[str], missing_docs: List[str]) -> Dict[str, Any]:
    profile = xlsx_profile.get("profile", {})
    flat_pairs = xlsx_profile.get("flat_pairs", {})
    records = xlsx_profile.get("records", [])
    clean_preview = xlsx_profile.get("clean_preview", {}) or {}
    worker_sheet_counts = xlsx_profile.get("worker_sheet_counts", {}) or {}
    worker_documents = [
        only_digits(
            record.get("documento")
            or record.get("numero_documento")
            or record.get("num_id_trabajador")
            or record.get("numero_de_identificacion")
            or ""
        )
        for record in records
    ]
    worker_documents = [value for value in worker_documents if value]
    employer_document = only_digits(profile.get("documento_empleador") or profile.get("nit") or flat_pairs.get("numerodocumentoempleador") or "")
    rejection_reasons: List[Dict[str, Any]] = []
    alerts: List[Dict[str, Any]] = []
    next_actions: List[str] = []
    row_errors: List[Dict[str, Any]] = []
    row_warnings: List[Dict[str, Any]] = []
    expected_workers = int(profile.get("numero_trabajadores") or 0) if str(profile.get("numero_trabajadores") or "").isdigit() else 0
    expected_sedes = int(profile.get("numero_sedes") or 0) if str(profile.get("numero_sedes") or "").isdigit() else 0
    parsed_sede_files = int(clean_preview.get("sede_files") or 0)
    workers_count_matches = expected_workers > 0 and len(records) == expected_workers

    missing_profile_fields: List[str] = []
    if not normalize_text(profile.get("empresa", "")):
        missing_profile_fields.append("empresa")
    if not only_digits(profile.get("nit", "")):
        missing_profile_fields.append("nit")
    if not only_digits(profile.get("documento", "")):
        missing_profile_fields.append("documento_representante")
    if not normalize_text(profile.get("tipo_afiliado", "")):
        missing_profile_fields.append("tipo_afiliado")
    if expected_workers <= 0:
        missing_profile_fields.append("numero_trabajadores")
    if expected_sedes <= 0:
        missing_profile_fields.append("numero_sedes")
    if missing_profile_fields:
        rejection_reasons.append(
            {
                "code": "XLSX_REQUIRED_FIELDS_MISSING",
                "severity": "blocker",
                "message": f"El XLSX no trae completos estos campos obligatorios: {', '.join(missing_profile_fields)}.",
            }
        )
        next_actions.append("Completar o corregir la cabecera del XLSX antes de continuar con la radicación.")

    if missing_docs:
        rejection_reasons.append(
            {
                "code": "MISSING_REQUIRED_DOCUMENTS",
                "severity": "blocker",
                "message": f"Faltan soportes obligatorios: {', '.join(missing_docs)}.",
            }
        )
        next_actions.append("Solicitar al cliente los soportes faltantes antes de continuar el trámite.")

    duplicates = sorted({doc for doc in worker_documents if worker_documents.count(doc) > 1})
    if duplicates:
        duplicate_details: List[str] = []
        for duplicate_doc in duplicates[:5]:
            appearances: List[str] = []
            for record in records:
                row_document = only_digits(
                    record.get("documento")
                    or record.get("numero_documento")
                    or record.get("num_id_trabajador")
                    or record.get("numero_de_identificacion")
                    or ""
                )
                if row_document != duplicate_doc:
                    continue
                full_name = normalize_text(
                    record.get("nombre")
                    or record.get("nombre_trabajador")
                    or " ".join(
                        value
                        for value in [
                            record.get("primer_nombre", ""),
                            record.get("segundo_nombre", ""),
                            record.get("primer_apellido", ""),
                            record.get("segundo_apellido", ""),
                        ]
                        if normalize_text(value)
                    )
                )
                sheet_name = normalize_text(record.get("_sheet", ""))
                row_number = normalize_text(record.get("_row", ""))
                parts = [duplicate_doc]
                if full_name:
                    parts.append(full_name)
                location = " · ".join(item for item in [sheet_name, f"fila {row_number}" if row_number else ""] if item)
                if location:
                    parts.append(location)
                appearances.append(" | ".join(parts))
            if appearances:
                duplicate_details.append("; ".join(appearances[:3]))
        rejection_reasons.append(
            {
                "code": "DUPLICATE_WORKER_DOCUMENTS",
                "severity": "blocker",
                "message": (
                    "El XLSX trae trabajadores duplicados por documento: "
                    + " ; ".join(duplicate_details)
                    + "."
                ),
            }
        )
        next_actions.append("Corregir en el XLSX los documentos repetidos indicados con hoja y fila antes de radicar.")

    if employer_document and employer_document in worker_documents:
        rejection_reasons.append(
            {
                "code": "WORKER_DOCUMENT_EQUALS_EMPLOYER",
                "severity": "blocker",
                "message": "El documento del trabajador no puede coincidir con el documento del empleador.",
            }
        )

    if expected_workers > 0 and not records:
        rows_message = (
            f"El XLSX declara {expected_workers} trabajador(es), pero no se pudieron leer filas individuales "
            "para validar cédulas repetidas, campos vacíos y consistencia fila a fila."
        )
        alerts.append(
            {
                "code": "XLSX_ROWS_NOT_PARSED",
                "severity": "alert",
                "message": rows_message,
            }
        )
        next_actions.append(
            f"Revisar la estructura interna de la hoja de trabajadores del XLSX; el archivo declara {expected_workers} trabajador(es) pero Imagine leyó 0 filas individuales."
        )

    empty_worker_sheets = [sheet for sheet, count in worker_sheet_counts.items() if int(count or 0) == 0]
    if empty_worker_sheets:
        message = "El XLSX tiene hojas de trabajadores sin registros válidos: " + ", ".join(empty_worker_sheets) + "."
        if workers_count_matches:
            alerts.append(
                {
                    "code": "EMPTY_WORKER_SHEETS",
                    "severity": "alert",
                    "message": message,
                }
            )
        else:
            rejection_reasons.append(
                {
                    "code": "EMPTY_WORKER_SHEETS",
                    "severity": "blocker",
                    "message": message,
                }
            )
            next_actions.append("Revisar las hojas de trabajadores que quedaron sin registros válidos o confirmar si deben eliminarse del contrato.")

    if expected_workers > 0 and records and len(records) != expected_workers:
        sheet_breakdown = ", ".join(
            f"{sheet}: {count}"
            for sheet, count in worker_sheet_counts.items()
        )
        rejection_reasons.append(
            {
                "code": "WORKER_COUNT_MISMATCH",
                "severity": "blocker",
                "message": (
                    f"El XLSX reporta {expected_workers} trabajadores, pero se leyeron {len(records)} filas de trabajadores."
                    + (f" Detalle por hoja: {sheet_breakdown}." if sheet_breakdown else "")
                ),
            }
        )
        next_actions.append("Revisar la hoja o sede donde el total de trabajadores no coincide con el consolidado del formulario.")

    if expected_sedes > 0 and parsed_sede_files > 0 and expected_sedes != parsed_sede_files:
        sheet_breakdown = ", ".join(
            f"{sheet}: {count}"
            for sheet, count in worker_sheet_counts.items()
        )
        message = (
            f"El contrato tiene {parsed_sede_files} sede(s) detectadas en el paquete, "
            f"pero solo {expected_sedes} sede(s) tienen trabajadores válidos en el XLSX."
            + (f" Detalle por hoja: {sheet_breakdown}." if sheet_breakdown else "")
        )
        if workers_count_matches:
            alerts.append(
                {
                    "code": "SEDES_COUNT_MISMATCH",
                    "severity": "alert",
                    "message": message,
                }
            )
        else:
            rejection_reasons.append(
                {
                    "code": "SEDES_COUNT_MISMATCH",
                    "severity": "blocker",
                    "message": message,
                }
            )

    valid_modalidades = {"PRESENCIAL", "TELETRABAJO", "CASA", "REMOTO"}
    valid_jornadas = {"UNICA", "ÚNICA", "TURNOS", "ROTATIVA"}
    valid_zonas = {"U", "R", "URBANA", "RURAL"}
    valid_sexos = {"M", "F", "MASCULINO", "FEMENINO"}
    for index, record in enumerate(records[:200], start=1):
        row_document = only_digits(
            record.get("documento")
            or record.get("numero_documento")
            or record.get("num_id_trabajador")
            or record.get("numero_de_identificacion")
            or ""
        )
        full_name = normalize_text(
            record.get("nombre")
            or record.get("nombre_trabajador")
            or " ".join(
                value
                for value in [
                    record.get("primer_nombre", ""),
                    record.get("segundo_nombre", ""),
                    record.get("primer_apellido", ""),
                    record.get("segundo_apellido", ""),
                ]
                if normalize_text(value)
            )
        )
        if not row_document:
            row_errors.append({"row": index, "code": "DOCUMENTO_VACIO", "message": f"Documento vacío en fila {index}.", "documento": ""})
        if not full_name:
            row_errors.append({"row": index, "code": "NOMBRE_VACIO", "message": f"Nombre vacío en fila {index}.", "documento": row_document})
        sexo = normalize_text(record.get("sexo", "") or record.get("sexo_identificacion", "")).upper()
        if sexo and sexo not in valid_sexos:
            row_errors.append({"row": index, "code": "SEXO_INVALIDO", "message": f"Sexo inválido en fila {index} ({sexo}).", "documento": row_document})

        zona_raw = record.get("zona", "") or record.get("zona_rural_urbana", "") or record.get("zona_(rural/urbana)", "")
        zona = normalize_haystack(zona_raw).upper()
        if zona and zona not in valid_zonas:
            row_errors.append({"row": index, "code": "ZONA_INVALIDA", "message": f"Zona inválida en fila {index} ({zona_raw}).", "documento": row_document})

        modalidad = normalize_haystack(record.get("modalidad", "")).upper()
        if modalidad and modalidad not in valid_modalidades:
            row_errors.append({"row": index, "code": "MODALIDAD_INVALIDA", "message": f"Modalidad inválida en fila {index} ({record.get('modalidad')}).", "documento": row_document})

        jornada = normalize_haystack(record.get("jornada", "")).upper()
        if jornada and jornada not in valid_jornadas:
            row_errors.append({"row": index, "code": "JORNADA_INVALIDA", "message": f"Jornada inválida en fila {index} ({record.get('jornada')}).", "documento": row_document})

        phone = only_digits(record.get("telefono") or "")
        if phone and len(phone) != 7:
            row_warnings.append({"row": index, "code": "TELEFONO_FORMATO", "message": f"Teléfono con longitud no legacy en fila {index}.", "documento": row_document})

        mobile = only_digits(record.get("celular") or "")
        if mobile and len(mobile) != 10:
            row_warnings.append({"row": index, "code": "CELULAR_FORMATO", "message": f"Celular con longitud no legacy en fila {index}.", "documento": row_document})

        email = normalize_text(record.get("mail") or record.get("correo") or record.get("correo_electronico") or "")
        if email and "@" not in email:
            row_warnings.append({"row": index, "code": "CORREO_FORMATO", "message": f"Correo con formato no válido en fila {index}.", "documento": row_document})

    if row_errors:
        rejection_reasons.extend(
            {"code": item["code"], "severity": "blocker", "message": item["message"]}
            for item in row_errors[:20]
        )
        next_actions.append("Corregir las filas inválidas del XLSX antes de continuar con la radicación.")

    listed_workers = _doc_by_type(docs, "listado_trabajadores")
    listed_workers_present = bool(listed_workers) or parsed_sede_files > 0 or any(doc.get("document_type") == "anexo_sedes" for doc in docs)
    if records and not listed_workers_present:
        alerts.append(
            {
                "code": "WORKER_LIST_NOT_FOUND",
                "severity": "alert",
                "message": "El XLSX contiene trabajadores, pero no se detectó un listado de trabajadores entre los adjuntos.",
            }
        )
        next_actions.append("Validar si el flujo exige listado de trabajadores o si la carga vino embebida solo en el XLSX.")

    if any(required in {"formulario_afiliacion", "anexo_sedes"} for required in required_docs):
        existing_types = {doc.get("document_type") for doc in docs}
        for expected_type in ["formulario_afiliacion", "anexo_sedes"]:
            if expected_type in required_docs and expected_type not in existing_types:
                alerts.append(
                    {
                        "code": f"OPTIONAL_{expected_type.upper()}_NOT_FOUND",
                        "severity": "alert",
                        "message": f"No se detectó {DOC_TYPE_LABELS.get(expected_type, expected_type)} en los adjuntos cargados.",
                    }
                )

    return {
        "approved": not rejection_reasons,
        "motivos_de_rechazo": rejection_reasons,
        "alerts": alerts,
        "row_errors": row_errors[:50],
        "row_warnings": row_warnings[:50],
        "next_actions": _unique_preserve(next_actions),
    }


def _build_validation_summary(xlsx_profile: Dict[str, Any], docs: List[Dict[str, Any]], missing_docs: List[str]) -> Dict[str, Any]:
    profile = xlsx_profile.get("profile", {})
    xlsx_document = only_digits(profile.get("documento", ""))
    xlsx_nit = _normalize_company_nit(profile.get("documento_empleador") or profile.get("nit", ""), docs)
    validations: List[Dict[str, Any]] = []
    alerts: List[Dict[str, Any]] = []
    matches: Dict[str, Any] = {}
    required_evidence = _infer_required_document_satisfaction(_build_required_documents(profile), docs, profile)

    cedula_docs = _doc_by_type(docs, "cedula")
    matched_cedula = None
    matched_cedula_candidate = ""
    if xlsx_document:
        for doc in cedula_docs:
            fields = doc.get("fields") or {}
            candidate = _best_numeric_candidate(xlsx_document, [fields.get("document_number", "")] + list(fields.get("all_numbers") or []))
            if candidate:
                matched_cedula = doc
                matched_cedula_candidate = candidate
                break
    inferred_cedula = required_evidence.get("cedula") or {}
    if not matched_cedula and inferred_cedula.get("satisfied") and inferred_cedula.get("filename"):
        matched_cedula = next((doc for doc in docs if doc.get("filename") == inferred_cedula.get("filename")), None)
        matched_cedula_candidate = inferred_cedula.get("matched", "")
    cedula_ok = bool(matched_cedula) if xlsx_document else bool(cedula_docs or inferred_cedula.get("satisfied"))
    validations.append(
        {
            "code": "CEDULA_MATCH_XLSX",
            "status": "OK" if cedula_ok else "ALERTA",
            "message": "La cedula coincide con el documento del XLSX." if cedula_ok else "La cedula OCR no coincide con el documento del XLSX.",
        }
    )
    if not cedula_ok:
        alerts.append(validations[-1])
    if xlsx_document:
        matches["cedula_principal"] = {
            "expected": xlsx_document,
            "matched": matched_cedula_candidate,
            "filename": (matched_cedula or {}).get("filename", ""),
            "ok": bool(matched_cedula),
        }

    rut_docs = _doc_by_type(docs, "rut")
    matched_rut = None
    rut_ok = False
    if xlsx_nit:
        matched_rut = next((doc for doc in rut_docs if _canonical_numeric_value(doc["fields"].get("nit", "")) == _canonical_numeric_value(xlsx_nit)), None)
        if not matched_rut:
            for doc in docs:
                fields = doc.get("fields") or {}
                preview = normalize_haystack(doc.get("text_preview", ""))
                candidates = [fields.get("nit", ""), fields.get("document_number", "")] + list(fields.get("all_numbers") or [])
                candidate = _best_numeric_candidate(xlsx_nit, candidates)
                if candidate and (
                    doc.get("document_type") in {"rut", "entrega_documentos", "camara_comercio", "soporte_ingresos"}
                    or "rut" in preview
                    or "dian" in preview
                    or "numero de identificacion tributaria" in preview
                ):
                    matched_rut = doc
                    break
        rut_ok = bool(matched_rut)
    else:
        rut_ok = bool(rut_docs)
    validations.append(
        {
            "code": "RUT_MATCH_XLSX",
            "status": "OK" if rut_ok else "ALERTA",
            "message": "El RUT coincide con el NIT del XLSX." if rut_ok else "El RUT no entrega un NIT consistente con el XLSX.",
        }
    )
    if not rut_ok:
        alerts.append(validations[-1])
    if xlsx_nit:
        matches["rut_nit"] = {
            "expected": xlsx_nit,
            "matched": _best_numeric_candidate(
                xlsx_nit,
                [
                    ((matched_rut or {}).get("fields", {}) or {}).get("nit", ""),
                    ((matched_rut or {}).get("fields", {}) or {}).get("document_number", ""),
                    *list((((matched_rut or {}).get("fields", {}) or {}).get("all_numbers") or [])),
                ],
            ),
            "filename": (matched_rut or {}).get("filename", ""),
            "ok": bool(matched_rut),
        }

    formulario_docs = _doc_by_type(docs, "formulario_afiliacion")
    representative_form = next((doc for doc in formulario_docs if only_digits((doc.get("fields") or {}).get("representative_document", "") or (doc.get("fields") or {}).get("document_number", ""))), None)
    form_doc = only_digits((representative_form or {}).get("fields", {}).get("representative_document", "") or (representative_form or {}).get("fields", {}).get("document_number", ""))
    profile_doc = only_digits(profile.get("documento", ""))
    if form_doc and profile_doc and not _numeric_similarity_match(form_doc, profile_doc):
        form_doc_candidates = [form_doc] + list(((representative_form or {}).get("fields", {}) or {}).get("all_numbers") or [])
        best_form_doc = _best_numeric_candidate(profile_doc, form_doc_candidates)
        if best_form_doc:
            form_doc = best_form_doc
        else:
            form_doc = profile_doc
    representative_cedula = None
    ced_doc = ""
    rep_doc_inferred = False
    if form_doc:
        for doc in cedula_docs:
            fields = doc.get("fields") or {}
            candidate = _best_numeric_candidate(form_doc, [fields.get("document_number", "")] + list(fields.get("all_numbers") or []))
            if candidate:
                representative_cedula = doc
                ced_doc = candidate
                break
    if not representative_cedula and inferred_cedula.get("satisfied") and form_doc and _numeric_similarity_match(form_doc, inferred_cedula.get("matched", "")):
        representative_cedula = next((doc for doc in docs if doc.get("filename") == inferred_cedula.get("filename")), None)
        ced_doc = inferred_cedula.get("matched", "")
        rep_doc_inferred = True
    rep_doc_ok = bool(form_doc and ced_doc and _numeric_similarity_match(form_doc, ced_doc))
    if representative_form or representative_cedula:
        if rep_doc_ok and rep_doc_inferred:
            rep_message = (
                f"La cédula del representante coincide por inferencia documental: formulario={form_doc or 'n/d'} "
                f"· cédula referenciada={ced_doc or 'n/d'}."
            )
        elif rep_doc_ok:
            rep_message = "La cédula del representante en el formulario coincide con la cédula adjunta."
        else:
            rep_message = f"La cédula del representante no coincide: formulario={form_doc or 'n/d'} · cédula adjunta={ced_doc or 'n/d'}."
        validations.append(
            {
                "code": "REPRESENTANTE_DOC_MATCH",
                "status": "OK" if rep_doc_ok else "ALERTA",
                "message": rep_message,
            }
        )
        matches["representante_documento"] = {
            "expected": form_doc,
            "matched": ced_doc,
            "formulario": (representative_form or {}).get("filename", ""),
            "cedula": (representative_cedula or {}).get("filename", ""),
            "inferred": rep_doc_inferred,
            "ok": rep_doc_ok,
        }
        if not rep_doc_ok:
            alerts.append(validations[-1])

    profile_company = normalize_haystack(profile.get("empresa", ""))
    profile_company_cmp = _normalize_company_compare(profile.get("empresa", ""))
    camara_docs = _doc_by_type(docs, "camara_comercio")
    camara_primary = camara_docs[0] if camara_docs else None
    formulario_primary = formulario_docs[0] if formulario_docs else None
    camara_company = normalize_haystack((camara_primary or {}).get("fields", {}).get("company_name", "") or (camara_primary or {}).get("text_preview", ""))
    form_company = normalize_haystack((formulario_primary or {}).get("fields", {}).get("company_name", "") or (formulario_primary or {}).get("text_preview", ""))
    camara_company_cmp = _normalize_company_compare((camara_primary or {}).get("fields", {}).get("company_name", "") or (camara_primary or {}).get("text_preview", ""))
    form_company_cmp = _normalize_company_compare((formulario_primary or {}).get("fields", {}).get("company_name", "") or (formulario_primary or {}).get("text_preview", ""))
    company_ok = False
    if camara_primary and formulario_primary:
        camara_tokens = [token for token in camara_company.split() if len(token) >= 5]
        profile_matches_camara = bool(profile_company_cmp and (profile_company_cmp in camara_company_cmp or camara_company_cmp in profile_company_cmp))
        profile_matches_form = bool(profile_company_cmp and form_company_cmp and (profile_company_cmp in form_company_cmp or form_company_cmp in profile_company_cmp))
        form_has_company = bool(form_company)
        company_ok = (
            profile_matches_camara
            or profile_matches_form
            or bool(camara_tokens and sum(1 for token in camara_tokens if token in form_company) >= min(2, len(camara_tokens)))
            or (not form_has_company and profile_matches_camara)
        )
        validations.append(
            {
                "code": "EMPRESA_MATCH_CAMARA_FORMULARIO",
                "status": "OK" if company_ok else "ALERTA",
                "message": "La razón social coincide entre cámara de comercio y formulario."
                if company_ok
                else "La razón social no coincide entre cámara de comercio y formulario.",
            }
        )
        matches["empresa_nombre"] = {
            "profile": profile.get("empresa", ""),
            "camara": (camara_primary or {}).get("fields", {}).get("company_name", ""),
            "formulario": (formulario_primary or {}).get("fields", {}).get("company_name", ""),
            "camara_file": (camara_primary or {}).get("filename", ""),
            "formulario_file": (formulario_primary or {}).get("filename", ""),
            "ok": company_ok,
        }
        if not company_ok:
            alerts.append(validations[-1])

    if "contrato" not in missing_docs:
        contrato_docs = _doc_by_type(docs, "contrato")
        contrato_ok = any(bool(doc["fields"].get("has_contrato_hint")) for doc in contrato_docs) or bool(contrato_docs)
        validations.append(
            {
                "code": "CONTRATO_PRESENTE",
                "status": "OK" if contrato_ok else "REVISAR",
                "message": "Se detecta soporte contractual." if contrato_ok else "El contrato existe pero requiere verificacion manual.",
            }
        )

    if "soporte_ingresos" not in missing_docs:
        ingresos_docs = _doc_by_type(docs, "soporte_ingresos")
        ingresos_ok = any(bool(doc["fields"].get("has_income_hint")) for doc in ingresos_docs) or bool(ingresos_docs)
        validations.append(
            {
                "code": "INGRESOS_PRESENTES",
                "status": "OK" if ingresos_ok else "REVISAR",
                "message": "Se detecta soporte de ingresos." if ingresos_ok else "El soporte de ingresos requiere confirmacion manual.",
            }
        )

    if camara_docs:
        recent_date = None
        recent_source = None
        for doc in camara_docs:
            fields = doc.get("fields") or {}
            parsed = None
            issue_date_textual = normalize_text(fields.get("issue_date_textual", ""))
            if issue_date_textual:
                try:
                    parsed = datetime.strptime(issue_date_textual, "%Y-%m-%d")
                except ValueError:
                    parsed = None
            if not parsed:
                parsed = _parse_spanish_date_text(doc.get("text_preview", "")) or _parse_date(doc.get("text_preview", ""))
            if parsed and (recent_date is None or parsed > recent_date):
                recent_date = parsed
                recent_source = doc.get("filename")
        if recent_date:
            age_days = (datetime.now() - recent_date).days
            ok = age_days <= 90
            issue_date_human = _format_date_es(recent_date)
            validations.append(
                {
                    "code": "CAMARA_VIGENTE",
                    "status": "OK" if ok else "ALERTA",
                    "message": (
                        f"Camara de comercio vigente. Fecha de expedicion: {issue_date_human}. Antigüedad: {age_days} dias."
                        if ok
                        else f"Camara de comercio vencida. Fecha de expedicion: {issue_date_human}. Antigüedad: {age_days} dias."
                    ),
                }
            )
            matches["camara_vigencia"] = {
                "issued_at": recent_date.strftime("%Y-%m-%d"),
                "issued_at_human": issue_date_human,
                "age_days": age_days,
                "filename": recent_source or "",
                "ok": ok,
            }
            if not ok:
                alerts.append(validations[-1])

    precheck = _build_precheck_summary(xlsx_profile, docs, _build_required_documents(profile), missing_docs)
    alerts.extend(item for item in precheck.get("alerts", []) if item not in alerts)
    precheck_reasons = list(precheck.get("motivos_de_rechazo", []))
    existing_reason_keys = {
        (item.get("code", ""), item.get("message", ""))
        for item in precheck_reasons
        if isinstance(item, dict)
    }
    for alert in alerts:
        if not isinstance(alert, dict):
            continue
        if normalize_haystack(alert.get("severity", "")).lower() == "alert":
            continue
        reason_key = (alert.get("code", ""), alert.get("message", ""))
        if reason_key in existing_reason_keys:
            continue
        precheck_reasons.append(
            {
                "code": alert.get("code", "VALIDATION_ALERT"),
                "severity": "blocker",
                "message": alert.get("message", ""),
            }
        )
        existing_reason_keys.add(reason_key)
    precheck["motivos_de_rechazo"] = precheck_reasons
    precheck["approved"] = not precheck_reasons
    return {
        "ok": precheck["approved"] and not missing_docs,
        "items": validations,
        "alerts": alerts,
        "matches": matches,
        "precheck": precheck,
    }


def _build_executive_report(label: str, xlsx_profile: Dict[str, Any], checklist: Dict[str, Any], decision: Dict[str, Any], validation_summary: Dict[str, Any]) -> Dict[str, Any]:
    profile = xlsx_profile.get("profile", {})
    employer_nit = _normalize_company_nit(profile.get("documento_empleador") or profile.get("nit", ""))
    records = xlsx_profile.get("records", []) or []
    worker_sheet_counts = xlsx_profile.get("worker_sheet_counts", {}) or {}
    raw_fecha_proceso = only_digits(profile.get("fecha_proceso") or "")[:8]
    if len(raw_fecha_proceso) != 8:
        raw_fecha_proceso = datetime.now().strftime("%Y%m%d")
    try:
        fecha_proceso_dt = datetime.strptime(raw_fecha_proceso, "%Y%m%d")
        fecha_proceso_human = _format_date_es(fecha_proceso_dt)
    except ValueError:
        fecha_proceso_human = raw_fecha_proceso
    derived_workers = profile.get("numero_trabajadores")
    derived_sedes = profile.get("numero_sedes")
    worker_count = int(derived_workers) if str(derived_workers).isdigit() else len(records)
    nomina_total = sum(
        _parse_nomina_value(
            record.get("salario")
            or record.get("salario_basico")
            or record.get("ibc")
            or record.get("ingreso_base_de_cotizacion")
            or ""
        )
        for record in records
    )
    active_worker_sedes = sum(1 for count in worker_sheet_counts.values() if int(count or 0) > 0)
    sedes_count = (
        active_worker_sedes
        if active_worker_sedes > 0
        else int(derived_sedes)
        if str(derived_sedes).isdigit()
        else sum(1 for item in (checklist.get("received_summary") or []) if item.get("document_type") == "anexo_sedes")
    )
    estado = "APROBADO" if decision.get("recommended_status") == "aprobable" else "RECHAZADO"
    errores: List[str] = []
    observaciones: List[str] = []
    for item in decision.get("blockers", []):
        message = item["message"] if isinstance(item, dict) else str(item)
        if message and message not in errores:
            errores.append(message)
    for item in validation_summary.get("alerts", []):
        message = item["message"] if isinstance(item, dict) else str(item)
        if message and message not in observaciones:
            observaciones.append(message)
    precheck = validation_summary.get("precheck", {})
    report_lines = [
        f"Reporte ejecutivo del caso {label}",
        f"Estado final: {estado}",
        f"Fecha de proceso: {fecha_proceso_human}",
        f"Afiliado: {profile.get('empresa', 'n/d')}",
        f"NIT: {employer_nit or 'n/d'}",
        f"Nómina total: {nomina_total}",
        f"Trabajadores: {worker_count}",
        f"Sedes: {sedes_count}",
    ]
    if checklist.get("missing"):
        report_lines.append(f"Faltantes: {', '.join(checklist.get('missing', []))}")
    if errores:
        report_lines.append("Hallazgos:")
        for item in errores[:8]:
            lines = format_reason_lines(item)
            if not lines:
                continue
            report_lines.append(f"- {lines[0]}")
            report_lines.extend(lines[1:])
    if observaciones:
        report_lines.append("Observaciones:")
        for item in observaciones[:8]:
            lines = format_reason_lines(item)
            if not lines:
                continue
            report_lines.append(f"- {lines[0]}")
            report_lines.extend(lines[1:])
    if precheck.get("next_actions"):
        report_lines.append("Acciones recomendadas:")
        report_lines.extend(f"- {item}" for item in precheck.get("next_actions", [])[:6])
    return {
        "estado_final": estado,
        "resumen_ejecutivo": {
            "caso": label,
            "empresa": profile.get("empresa", ""),
            "nit": employer_nit,
            "documento": profile.get("documento", ""),
            "fecha_proceso": raw_fecha_proceso,
            "fecha_proceso_human": fecha_proceso_human,
            "nomina_total": nomina_total,
            "numero_trabajadores": worker_count,
            "numero_sedes": sedes_count,
            "estado": estado,
            "faltantes": checklist.get("missing", []),
            "errores": errores,
            "observaciones": observaciones,
            "matches": validation_summary.get("matches", {}),
            "siguiente_paso": decision.get("next_step", ""),
            "acciones_recomendadas": precheck.get("next_actions", []),
        },
        "texto": "\n".join(report_lines),
    }


def _build_926_draft(profile: Dict[str, Any], checklist: Dict[str, Any], decision: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if decision.get("recommended_status") != "aprobable":
        return None
    documento = only_digits(profile.get("documento", ""))
    nit = only_digits(profile.get("nit", ""))
    empresa = normalize_text(profile.get("empresa", ""))
    afiliado = normalize_text(profile.get("tipo_afiliado", ""))
    lines = [
        f"1|{nit}|{empresa[:60]}",
        f"2|{documento}|{afiliado[:30]}|OK",
        f"9|{','.join(checklist.get('received', []))}|READY",
    ]
    return {
        "filename": f"borrador_926_{slugify(empresa or 'caso')}.txt",
        "content": "\n".join(lines) + "\n",
        "note": "Borrador operativo inicial. Aun no reemplaza el generador 926 legacy completo.",
    }


def _build_926_output(case_id: str, xlsx_profile: Dict[str, Any], checklist: Dict[str, Any], decision: Dict[str, Any]) -> Dict[str, Any]:
    if decision.get("recommended_status") != "aprobable":
        return {
            "available": False,
            "mode": "blocked",
            "reason": "El caso no está aprobable; el 926 permanece bloqueado.",
            "draft": None,
            "legacy": None,
        }

    profile = xlsx_profile.get("profile", {})
    lote = normalize_text(profile.get("lote") or profile.get("idtramite") or "")
    legacy_result = {"available": False, "ok": False, "error": "Sin lote para bridge legacy."}
    if lote:
        legacy_result = generate_legacy_flatfile_926_http(lote=lote)
        if not legacy_result.get("ok"):
            state_result = generate_legacy_flatfile_926(lote=lote)
            if state_result.get("ok"):
                legacy_result = state_result
    draft = _build_926_draft(profile, checklist, decision)
    if legacy_result.get("ok"):
        return {
            "available": True,
            "mode": "legacy",
            "reason": "",
            "draft": draft,
            "legacy": legacy_result,
        }
    return {
        "available": bool(draft),
        "mode": "draft",
        "reason": legacy_result.get("error", "Bridge legacy no disponible; se entrega borrador."),
        "draft": draft,
        "legacy": legacy_result,
    }


def _legacy_post(path: str, payload: Dict[str, Any], timeout: float = 120.0) -> Dict[str, Any]:
    base_url = str(settings.legacy_backend_url or "").strip().rstrip("/")
    if not base_url:
        raise RuntimeError("legacy_backend_url no configurado.")
    url = f"{base_url}/{path.lstrip('/')}"
    response = httpx.post(url, json=payload, timeout=timeout)
    response.raise_for_status()
    return response.json()


def _legacy_build_926_http(
    lote: str,
    base: str = "temporal",
    strict_validate: bool = True,
    fecha_proceso: str = "",
    lote_usuario: str = "",
) -> Dict[str, Any]:
    base_url = str(settings.legacy_backend_url or "").strip().rstrip("/")
    if not base_url:
        raise RuntimeError("legacy_backend_url no configurado.")
    response = httpx.get(
        f"{base_url}/legacy/flatfile/build",
        params={
            "lote": lote,
            "from_db": "true",
            "base": base,
            "strict_validate": "true" if strict_validate else "false",
            "fecha_proceso": only_digits(fecha_proceso)[:8],
            "lote_usuario": only_digits(lote_usuario),
        },
        timeout=120.0,
    )
    if response.status_code != 200:
        raise RuntimeError(f"HTTP {response.status_code}: {response.text[:300]}")
    return {
        "filename": f"BkCargue_{lote or 'generated'}.txt",
        "content": response.text,
        "headers": dict(response.headers),
    }


def _get_xlsx_file_entry(payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    for item in payload.get("files", []):
        filename = str(item.get("filename", "")).lower()
        if filename.endswith((".xlsx", ".xlsm", ".xls")):
            return item
    return None


def _build_manifest_step(payload: Dict[str, Any], analysis: Dict[str, Any]) -> Dict[str, Any]:
    profile = (analysis.get("xlsx_profile") or {}).get("profile", {})
    lote = normalize_text(profile.get("lote") or profile.get("idtramite") or payload.get("id") or "")
    lines = [
        f"CASE={payload.get('id')}",
        f"LOTE={lote or slugify(profile.get('empresa') or payload.get('label') or 'lote')}",
        f"EMPRESA={normalize_text(profile.get('empresa', ''))}",
        f"NIT={only_digits(profile.get('nit', ''))}",
        f"TIPO_AFILIADO={normalize_text(profile.get('tipo_afiliado', ''))}",
        f"DOCUMENTO={only_digits(profile.get('documento', ''))}",
    ]
    return {
        "filename": f"{slugify(lote or payload.get('label') or payload.get('id') or 'lote')}.txt",
        "content": "\n".join(lines).strip() + "\n",
        "lote": lote or slugify(profile.get("empresa") or payload.get("label") or "lote"),
    }


def _build_contrato_clean(xlsx_profile: Dict[str, Any], docs: List[Dict[str, Any]]) -> Dict[str, Any]:
    profile = xlsx_profile.get("profile", {})
    flat_pairs = xlsx_profile.get("flat_pairs", {})
    employer_doc = only_digits(profile.get("nit") or profile.get("documento_empleador") or "")
    employer_name = normalize_text(profile.get("empresa") or "EMPRESA EN PROCESO")
    worker_name = normalize_text(profile.get("nombre") or "")
    worker_doc = only_digits(profile.get("documento") or "")
    rep_name = normalize_text(flat_pairs.get("nombrerepresentantelegal") or worker_name or employer_name)
    rep_doc = only_digits(flat_pairs.get("numerodocumentorepresnetantelegal") or worker_doc)
    inicio = flat_pairs.get("iniciocobertura") or datetime.now().strftime("%Y-%m-%d")
    radicacion = flat_pairs.get("fecharadicacion") or datetime.now().strftime("%Y-%m-%d")
    actividad = only_digits(flat_pairs.get("actividadeconomicaempleador") or "5861001") or "5861001"
    tipo_aportante = normalize_text(flat_pairs.get("tipoaportante") or "05")
    naturaleza = normalize_text(flat_pairs.get("naturalezajuridica") or "Privada")
    ciudad = normalize_text(flat_pairs.get("ciudadempleador") or "11001")
    localidad = normalize_text(flat_pairs.get("localidadempleador") or "NA")
    zona = normalize_text(flat_pairs.get("zonaempleador") or "U")
    direccion = normalize_text(flat_pairs.get("direccionempleador") or "DIRECCION PENDIENTE")
    telefono = only_digits(flat_pairs.get("telefonoprincipalempleador") or "6010000000")
    email = normalize_text(flat_pairs.get("correoelectronicoempleador") or "contacto@empresa.test").lower()
    contrato_num = only_digits(flat_pairs.get("numerocontrato") or employer_doc or worker_doc or "1001")

    lines = [
        f"1. Apellidos y nombres o razón social|{employer_name}|2. Tipo de documento|NI|3. Número de documento o NIT|{employer_doc}",
        f"1. Tipo de trámite|X|2. Naturaleza jurídica del empleador|{naturaleza}|3. Tipo de aportante|{tipo_aportante}",
        f"4. Apellidos y nombres del Representante Legal|{rep_name}",
        f"5. Tipo de documento|CC|6. Número de documento|{rep_doc}|7. Correo electrónico|{email}",
        f"1. Datos de la sede principal|Dirección de la sede principal|{direccion}|Teléfono fijo/celular|{telefono}",
        f"|1|PRINCIPAL|Correo electrónico|{email}",
        f"Municipio/Distrito|{ciudad}|Zona|{zona}|Localidad/Comuna|{localidad}|Departamento|BOGOTA D.C.",
        f"1. ARL de la cual se traslada|10|2. Clase de riesgo|I|Actividad económica|{actividad}",
        f"{radicacion}T00:00:00|{inicio}T00:00:00|{contrato_num}|01",
    ]
    return {
        "filename": "contrato_clean_auto.txt",
        "content": "\n".join(lines) + "\n",
        "lines": len(lines),
    }


def _build_independientes_clean(xlsx_profile: Dict[str, Any]) -> Dict[str, Any]:
    profile = xlsx_profile.get("profile", {})
    flat_pairs = xlsx_profile.get("flat_pairs", {})
    worker_doc = only_digits(profile.get("documento") or "")
    full_name = normalize_text(profile.get("nombre") or "TRABAJADOR INDEPENDIENTE")
    parts = full_name.split()
    primer_nombre = parts[0] if parts else "TRABAJADOR"
    segundo_nombre = parts[1] if len(parts) > 1 else ""
    primer_apellido = parts[2] if len(parts) > 2 else "INDEPENDIENTE"
    segundo_apellido = parts[3] if len(parts) > 3 else ""
    inicio = flat_pairs.get("iniciocobertura") or datetime.now().strftime("%Y-%m-%d")
    fin = flat_pairs.get("finalizacioncontrato") or inicio
    ibc = only_digits(flat_pairs.get("ibc") or flat_pairs.get("ingresomensual") or "1423500") or "1423500"
    actividad = only_digits(flat_pairs.get("actividadeconomicaempleador") or "5861001") or "5861001"
    fecha_nacimiento = only_digits(flat_pairs.get("fechanacimiento") or "19900101") or "19900101"
    inicio_legacy = only_digits(inicio) or datetime.now().strftime("%Y%m%d")
    fin_legacy = only_digits(fin) or inicio_legacy
    departamento = only_digits(flat_pairs.get("departamento") or "11") or "11"
    municipio = only_digits(flat_pairs.get("ciudad") or flat_pairs.get("municipio") or "11001") or "11001"
    zona = normalize_text(flat_pairs.get("zona") or "U").upper()[:1] or "U"
    localidad = normalize_text(flat_pairs.get("localidad") or "LOCALIDAD")
    direccion = normalize_text(flat_pairs.get("direccion") or "DIRECCION RESIDENCIA")
    telefono = only_digits(flat_pairs.get("telefono") or "6010000") or "6010000"
    celular = only_digits(flat_pairs.get("celular") or "3000000000") or "3000000000"
    correo = normalize_text(flat_pairs.get("correo") or "trabajador@test.com").lower()
    eps = normalize_text(flat_pairs.get("eps") or "EPS DEMO").upper()
    codigo_eps = only_digits(flat_pairs.get("codigo_eps") or "00001") or "00001"
    afp = normalize_text(flat_pairs.get("afp") or "AFP DEMO").upper()
    codigo_afp = only_digits(flat_pairs.get("codigo_afp") or "00002") or "00002"
    tipo_cotizante = only_digits(flat_pairs.get("tipo_cotizante") or "51") or "51"
    subtipo_cotizante = only_digits(flat_pairs.get("subtipo_cotizante") or "0") or "0"
    tipo_contrato = normalize_text(flat_pairs.get("tipo_contrato") or "1")
    valor_contrato = only_digits(flat_pairs.get("valor_contrato") or ibc) or ibc
    nombre_actividad = normalize_text(flat_pairs.get("nombre_actividad") or "ACTIVIDAD ECONOMICA")
    clase_riesgo = normalize_text(flat_pairs.get("clase_riesgo") or "1")
    tasa_riesgo = normalize_text(flat_pairs.get("tasa_riesgo") or "0.522")
    lote = normalize_text(profile.get("lote") or profile.get("idtramite") or "AUTO")
    row = {field: "" for field in LEGACY_INDEPENDIENTES_FIELDS}
    row.update(
        {
            "sr": "2",
            "linea": "1",
            "tipodocumento": "CC",
            "documento": worker_doc,
            "primer_apellido": primer_apellido.upper(),
            "segundo_apellido": segundo_apellido.upper(),
            "primer_nombre": primer_nombre.upper(),
            "segundo_nombre": segundo_nombre.upper(),
            "fecha_nacimiento": fecha_nacimiento,
            "sexo": "M",
            "direccion": direccion.upper(),
            "departamento": departamento,
            "municipio": municipio,
            "zona": zona,
            "localidad": localidad.upper(),
            "telefono": telefono,
            "celular": celular,
            "correo": correo,
            "eps": eps,
            "codigo_eps": codigo_eps,
            "afp": afp,
            "codigo_afp": codigo_afp,
            "tipo_cotizante": tipo_cotizante,
            "subtipo_cotizante": subtipo_cotizante,
            "modalidad": "PRESENCIAL",
            "tipo_contrato": tipo_contrato,
            "transporte": "NO",
            "fecha_inicio_contrato": inicio_legacy,
            "fecha_fin_contrato": fin_legacy,
            "meses_contrato": "1",
            "valor_contrato": valor_contrato,
            "valor_mensual": ibc,
            "ibc": ibc,
            "actividad_economica": actividad,
            "nombre_actividad": nombre_actividad.upper(),
            "clase_riesgo": clase_riesgo,
            "tasa_riesgo": tasa_riesgo,
            "lunes": "X",
            "martes": "X",
            "miercoles": "X",
            "jueves": "X",
            "viernes": "X",
            "codigo_ct": "000001",
            "nombre_ct": "PRINCIPAL",
            "actividad_economica_ct": actividad,
            "clase_riesgo_ct": clase_riesgo,
            "tasa_riesgo_ct": tasa_riesgo,
            "direccion_ct": "DIRECCION CT",
            "departamento_ct": departamento,
            "ciudad_ct": municipio,
            "zona_ct": zona,
            "telefono_ct": telefono,
            "celular_ct": celular,
            "correo_ct": correo,
            "localidad_ct": localidad.upper(),
            "lote": lote,
            "tipo_salario": "VARIABLE",
        }
    )
    lines = ["!".join(str(row[field]) for field in LEGACY_INDEPENDIENTES_FIELDS)]
    return {"filename": "independientes_clean_auto.txt", "content": "\n".join(lines) + "\n", "lines": 1}


def _workflow_step(name: str, title: str, status: str, detail: str = "", payload: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    return {
        "name": name,
        "title": title,
        "status": status,
        "detail": detail,
        "payload": payload or {},
        "at": utc_now(),
    }


def run_case_workflow(case_id: str) -> Dict[str, Any]:
    payload = analyze_case(case_id)
    analysis = payload.get("analysis") or {}
    profile = (analysis.get("xlsx_profile") or {}).get("profile", {})
    xlsx_entry = _get_xlsx_file_entry(payload)
    xlsx_bytes = Path(xlsx_entry["stored_path"]).read_bytes() if xlsx_entry else b""
    xlsx_b64 = base64.b64encode(xlsx_bytes).decode("ascii") if xlsx_bytes else ""
    flat_pairs = ((analysis.get("xlsx_profile") or {}).get("flat_pairs", {}) or {})
    lote = normalize_text(profile.get("lote") or profile.get("idtramite") or "")
    idtramite = only_digits(profile.get("idtramite") or flat_pairs.get("idtramite") or "")
    legacy_lote_usuario = only_digits(
        profile.get("lote_usuario") or flat_pairs.get("lote_usuario") or flat_pairs.get("lt_usuario") or payload.get("lote_usuario") or ""
    )
    legacy_fecha_proceso = only_digits(
        profile.get("fecha_proceso") or flat_pairs.get("fecha_proceso") or payload.get("fecha_proceso") or ""
    )[:8]
    if len(legacy_fecha_proceso) != 8:
        legacy_fecha_proceso = datetime.now().strftime("%Y%m%d")
    base = "temporal"
    timeline: List[Dict[str, Any]] = []
    final_report = analysis.get("reporte_ejecutivo")
    output_926 = analysis.get("output_926")

    timeline.append(
        _workflow_step(
            "prevalidacion_documental",
            "Prevalidación documental",
            "ok" if analysis.get("decision", {}).get("recommended_status") == "aprobable" else "blocked",
            analysis.get("decision", {}).get("summary", ""),
            {"decision": analysis.get("decision"), "checklist": analysis.get("checklist")},
        )
    )
    if analysis.get("decision", {}).get("recommended_status") != "aprobable":
        workflow = {
            "status": "stopped_prevalidacion",
            "current_step": "prevalidacion_documental",
            "steps": timeline,
            "stop_reason": "La prevalidación documental no fue aprobada.",
            "executive_report_precheck": analysis.get("reporte_ejecutivo"),
            "executive_report_final": None,
            "output_926": output_926,
        }
        payload["analysis"]["workflow_run"] = workflow
        payload["updated_at"] = utc_now()
        save_case(payload)
        return payload

    manifest = _build_manifest_step(payload, analysis)
    lote = lote or manifest["lote"]
    idtramite = idtramite or str(int(datetime.now().timestamp()))[-8:]
    if not legacy_lote_usuario:
        legacy_lote_usuario = build_generated_lote_usuario(legacy_fecha_proceso)
    payload["lote_usuario"] = legacy_lote_usuario
    payload["fecha_proceso"] = legacy_fecha_proceso
    analysis.setdefault("xlsx_profile", {}).setdefault("profile", {})
    analysis["xlsx_profile"]["profile"]["lote_usuario"] = legacy_lote_usuario
    analysis["xlsx_profile"]["profile"]["fecha_proceso"] = legacy_fecha_proceso
    timeline.append(
        _workflow_step(
            "manifesto_paso1",
            "Generación de manifiesto",
            "ok",
            f"Manifiesto {manifest['filename']} listo para lote {lote}.",
            {"manifest": manifest},
        )
    )

    try:
        bootstrap_out = _legacy_post(
            "ruta-inclusion/bootstrap-tramite",
            {
                "base": base,
                "idtramite": idtramite,
                "lote": lote,
                "estado": "Estudio",
                "usuario": "nova_case_workflow",
                "lote_usuario": legacy_lote_usuario,
                "fecha_proceso": legacy_fecha_proceso,
            },
            timeout=60.0,
        )
        upload_out = _legacy_post("legacy/ls/upload-archivo-txt", {"fileid": lote, "content": manifest["content"]}, timeout=60.0)
        read_out = _legacy_post("legacy/opc", {"sw": "LS", "sw1": "archivo_txt", "fileid": lote}, timeout=60.0)
        deliver_out = _legacy_post(
            "legacy/opc",
            {
                "sw": "LS",
                "sw1": "fileid-data",
                "fileid": lote,
                "var1": "CC",
                "ol": "1",
                "userid": "nova_case_workflow",
                "fechabd": datetime.now().strftime("%Y%m%d"),
                "fechashora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
            timeout=120.0,
        )
        timeline.append(
            _workflow_step(
                "carga_lote_legacy",
                "Carga de lote TXT",
                "ok",
                f"Lote {lote} e idtrámite {idtramite} entregados al clone legacy.",
                {"bootstrap": bootstrap_out, "upload": upload_out, "read": read_out, "deliver": deliver_out},
            )
        )
    except Exception as exc:
        timeline.append(_workflow_step("carga_lote_legacy", "Carga de lote TXT", "failed", str(exc)))
        workflow = {
            "status": "stopped_lote",
            "current_step": "carga_lote_legacy",
            "steps": timeline,
            "stop_reason": f"No pude cargar el lote al clone legacy: {exc}",
            "executive_report_precheck": analysis.get("reporte_ejecutivo"),
            "executive_report_final": None,
            "output_926": output_926,
        }
        payload["analysis"]["workflow_run"] = workflow
        payload["updated_at"] = utc_now()
        save_case(payload)
        return payload

    timeline.append(
        _workflow_step(
            "limpieza_preparacion",
            "Limpieza y preparación",
            "ok",
            "El expediente quedó limpio y clasificado para validaciones.",
            {"documents": analysis.get("documents", [])},
        )
    )

    clean_output: Dict[str, Any] = {}
    if xlsx_entry and xlsx_bytes:
        clean_output = _generate_clean_via_legacy_nova(xlsx_entry["filename"], xlsx_bytes)
        if not bool(clean_output.get("ok")):
            workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            clean_output = _generate_clean_from_workbook(workbook, xlsx_entry["filename"])
            clean_output["_source"] = "local_fallback"

    contrato_clean = clean_output.get("contrato_clean") if isinstance(clean_output.get("contrato_clean"), dict) else None
    trabajadores_clean_multi = clean_output.get("trabajadores_clean_multi") if isinstance(clean_output.get("trabajadores_clean_multi"), list) else []
    independientes_clean = clean_output.get("independientes_clean") if isinstance(clean_output.get("independientes_clean"), dict) else None
    if (not contrato_clean or not str(contrato_clean.get("content") or "").strip()) and xlsx_entry:
        contrato_clean = _build_contrato_clean(analysis.get("xlsx_profile") or {}, analysis.get("documents") or [])
    if not trabajadores_clean_multi and xlsx_entry:
        fallback_indep = _build_independientes_clean(analysis.get("xlsx_profile") or {})
        if str(fallback_indep.get("content") or "").strip():
            independientes_clean = fallback_indep

    contract_fields = _extract_employer_from_contract_text(str((contrato_clean or {}).get("content") or ""))
    if contract_fields:
        company_name = normalize_text(contract_fields.get("empresa", ""))
        nit_value = only_digits(contract_fields.get("nit", ""))
        rep_name = normalize_text(contract_fields.get("representante_legal", ""))
        rep_doc = only_digits(contract_fields.get("doc_representante", ""))
        analysis.setdefault("xlsx_profile", {}).setdefault("profile", {})
        if company_name and len(company_name) >= 5:
            analysis["xlsx_profile"]["profile"]["empresa"] = profile.get("empresa") or company_name
        if 8 <= len(nit_value) <= 12:
            employer_document_hint = only_digits(profile.get("documento_empleador") or nit_value)
            analysis["xlsx_profile"]["profile"]["nit"] = employer_document_hint or nit_value
            analysis["xlsx_profile"]["profile"]["documento_empleador"] = employer_document_hint or nit_value
        if rep_name and len(rep_name) >= 5:
            analysis["xlsx_profile"]["profile"]["nombre"] = profile.get("nombre") or rep_name
        if 6 <= len(rep_doc) <= 15:
            analysis["xlsx_profile"]["profile"]["documento"] = profile.get("documento") or rep_doc
        profile = analysis["xlsx_profile"]["profile"]

    try:
        empleador_out = _legacy_post(
            "ruta-inclusion/importar-empleador-contrato",
            {
                "base": base,
                "idtramite": idtramite,
                "content": contrato_clean["content"],
                "strict_validate": True,
                "usuario": "nova_case_workflow",
            },
            timeout=120.0,
        )
        sedes_out: List[Dict[str, Any]] = []
        trabajadores_out: List[Dict[str, Any]] = []
        for index, sede_clean in enumerate(trabajadores_clean_multi):
            replace = index == 0
            sede_result = _legacy_post(
                "ruta-inclusion/importar-sede-contrato",
                {
                    "base": base,
                    "idtramite": idtramite,
                    "content": sede_clean["content"],
                    "strict_validate": False,
                    "replace_existing": replace,
                    "auto_skip_empty_clean": True,
                    "usuario": "nova_case_workflow",
                },
                timeout=120.0,
            )
            sedes_out.append({"file": sede_clean.get("filename"), "result": sede_result})
            if bool(sede_result.get("skipped")):
                continue
            sed_sr = int(sede_result.get("sr_inserted") or 0)
            worker_payload = {
                "base": base,
                "idtramite": idtramite,
                "content": sede_clean["content"],
                "contrato_content": contrato_clean["content"],
                "strict_validate": True,
                "replace_existing": replace,
                "allow_empty": True,
                "usuario": "nova_case_workflow",
            }
            if sed_sr > 0:
                worker_payload["sr_override"] = sed_sr
            worker_result = _legacy_post("ruta-inclusion/importar-trabajadores-contrato", worker_payload, timeout=120.0)
            trabajadores_out.append({"file": sede_clean.get("filename"), "result": worker_result})

        indep_out: List[Dict[str, Any]] = []
        if independientes_clean and str(independientes_clean.get("content") or "").strip():
            indep_result = _legacy_post(
                "ruta-inclusion/importar-trabajadores-contrato",
                {
                    "base": base,
                    "idtramite": idtramite,
                    "content": independientes_clean["content"],
                    "contrato_content": contrato_clean["content"],
                    "strict_validate": True,
                    "replace_existing": not trabajadores_clean_multi,
                    "allow_empty": True,
                    "source_kind": "independientes_legacy",
                    "usuario": "nova_case_workflow",
                },
                timeout=120.0,
            )
            indep_out.append({"file": independientes_clean.get("filename"), "result": indep_result})

        comisiones_out: List[Dict[str, Any]] = []
        for doc in analysis.get("documents", []):
            if doc.get("document_type") not in {"entrega_documentos", "comision"}:
                continue
            content = str(doc.get("ocr_text") or doc.get("text_preview") or "").strip()
            haystack = normalize_haystack(content)
            if "reconocimiento variable integral" not in haystack and "participacion" not in haystack and "porcentaje" not in haystack:
                continue
            try:
                com_result = _legacy_post(
                    "ruta-inclusion/importar-comisiones-contrato",
                    {
                        "base": base,
                        "idtramite": idtramite,
                        "content": content,
                        "replace_existing": True,
                        "usuario": "nova_case_workflow",
                    },
                    timeout=120.0,
                )
                comisiones_out.append({"file": doc.get("filename"), "result": com_result})
                if bool(com_result.get("rows_inserted", 0)):
                    break
            except Exception as com_exc:
                comisiones_out.append({"file": doc.get("filename"), "error": str(com_exc)})

        timeline.append(
            _workflow_step(
                "importacion_proc",
                "Importación a proc_servicios",
                "ok",
                f"Se importó empleador, {len(sedes_out)} sede(s) y trabajadores para idtrámite {idtramite}.",
                {
                    "idtramite": idtramite,
                    "clean_source": clean_output.get("_source", "local"),
                    "contrato_clean": contrato_clean,
                    "trabajadores_clean_multi": [
                        {"filename": item.get("filename"), "lines": item.get("lines")}
                        for item in trabajadores_clean_multi
                    ],
                    "independientes_clean": (
                        {"filename": independientes_clean.get("filename"), "lines": independientes_clean.get("lines")}
                        if independientes_clean
                        else None
                    ),
                    "empleador": empleador_out,
                    "sedes": sedes_out,
                    "trabajadores": trabajadores_out,
                    "independientes": indep_out,
                    "comisiones": comisiones_out,
                },
            )
        )
    except Exception as exc:
        timeline.append(_workflow_step("importacion_proc", "Importación a proc_servicios", "failed", str(exc)))
        workflow = {
            "status": "stopped_importacion",
            "current_step": "importacion_proc",
            "steps": timeline,
            "stop_reason": f"No pude importar el caso a proc_servicios: {exc}",
            "executive_report_precheck": analysis.get("reporte_ejecutivo"),
            "executive_report_final": None,
            "output_926": output_926,
        }
        payload["analysis"]["workflow_run"] = workflow
        payload["updated_at"] = utc_now()
        save_case(payload)
        return payload

    try:
        import_proc_out = _legacy_post(
            "legacy/db/import-proc-servicios",
            {
                "base": base,
                "lote": lote,
                "estado": "Estudio",
                "idtramite": idtramite,
                "limit": 200,
                "apply_to_db": True,
            },
            timeout=120.0,
        )
        timeline.append(
            _workflow_step(
                "sync_engine",
                "Sincronización al engine legacy",
                "ok",
                f"El engine legacy quedó sincronizado para lote {lote}.",
                {"import": import_proc_out},
            )
        )
    except Exception as exc:
        timeline.append(_workflow_step("sync_engine", "Sincronización al engine legacy", "failed", str(exc)))
        workflow = {
            "status": "stopped_sync_engine",
            "current_step": "sync_engine",
            "steps": timeline,
            "stop_reason": f"No pude sincronizar proc_servicios al engine legacy: {exc}",
            "executive_report_precheck": analysis.get("reporte_ejecutivo"),
            "executive_report_final": None,
            "output_926": output_926,
        }
        payload["analysis"]["workflow_run"] = workflow
        payload["updated_at"] = utc_now()
        save_case(payload)
        return payload

    pre_payload: Dict[str, Any] = {"lote": lote, "from_db": True, "base": "temporal"}
    if xlsx_b64:
        pre_payload["excel_file_base64"] = xlsx_b64
    try:
        prebuild_out = _legacy_post("legacy/rules/prebuild-check", pre_payload, timeout=120.0)
        prebuild_ok = bool(prebuild_out.get("ok", False))
        timeline.append(
            _workflow_step(
                "prebuild_validaciones",
                "Prebuild y validaciones",
                "ok" if prebuild_ok else "blocked",
                "Prebuild legacy aprobado." if prebuild_ok else "Prebuild legacy bloqueó el flujo.",
                {"prebuild": prebuild_out},
            )
        )
        if not prebuild_ok:
            workflow = {
                "status": "stopped_prebuild",
                "current_step": "prebuild_validaciones",
                "steps": timeline,
                "stop_reason": "El prebuild legacy no fue aprobado.",
                "executive_report_precheck": analysis.get("reporte_ejecutivo"),
                "executive_report_final": None,
                "output_926": output_926,
            }
            payload["analysis"]["workflow_run"] = workflow
            payload["updated_at"] = utc_now()
            save_case(payload)
            return payload
    except Exception as exc:
        timeline.append(_workflow_step("prebuild_validaciones", "Prebuild y validaciones", "failed", str(exc)))
        workflow = {
            "status": "stopped_prebuild",
            "current_step": "prebuild_validaciones",
            "steps": timeline,
            "stop_reason": f"No pude ejecutar prebuild legacy: {exc}",
            "executive_report_precheck": analysis.get("reporte_ejecutivo"),
            "executive_report_final": None,
            "output_926": output_926,
        }
        payload["analysis"]["workflow_run"] = workflow
        payload["updated_at"] = utc_now()
        save_case(payload)
        return payload

    try:
        pre_report_payload: Dict[str, Any] = {
            "lote": lote,
            "base": base,
            "from_db": True,
            "generate_926_if_missing": False,
        }
        if xlsx_b64:
            pre_report_payload["excel_file_base64"] = xlsx_b64
        pre_report = _legacy_post("legacy/reporte-ejecutivo-contrato", pre_report_payload, timeout=120.0)
        timeline.append(
            _workflow_step(
                "reporte_previo",
                "Reporte ejecutivo previo",
                "ok",
                "Reporte ejecutivo previo al 926 generado.",
                {"report": pre_report},
            )
        )
    except Exception as exc:
        pre_report = None
        timeline.append(_workflow_step("reporte_previo", "Reporte ejecutivo previo", "failed", str(exc)))

    try:
        generated_926 = _legacy_build_926_http(
            lote=lote,
            base=base,
            strict_validate=True,
            fecha_proceso=legacy_fecha_proceso,
            lote_usuario=legacy_lote_usuario,
        )
        output_926 = {
            "available": True,
            "mode": "legacy",
            "reason": "",
            "draft": analysis.get("draft_926"),
            "legacy": {"available": True, "ok": True, **generated_926},
        }
        timeline.append(
            _workflow_step(
                "generacion_926",
                "Generación 926",
                "ok",
                f"Archivo 926 generado para lote {lote}.",
                {"filename": generated_926.get("filename")},
            )
        )
    except Exception as exc:
        timeline.append(_workflow_step("generacion_926", "Generación 926", "failed", str(exc)))
        workflow = {
            "status": "stopped_926",
            "current_step": "generacion_926",
            "steps": timeline,
            "stop_reason": f"No pude generar el 926 legacy: {exc}",
            "executive_report_precheck": analysis.get("reporte_ejecutivo"),
            "executive_report_final": pre_report,
            "output_926": analysis.get("output_926"),
        }
        payload["analysis"]["workflow_run"] = workflow
        payload["updated_at"] = utc_now()
        save_case(payload)
        return payload

    try:
        final_report_payload: Dict[str, Any] = {
            "lote": lote,
            "base": base,
            "from_db": True,
            "generate_926_if_missing": True,
        }
        if xlsx_b64:
            final_report_payload["excel_file_base64"] = xlsx_b64
        final_report = _legacy_post("legacy/reporte-ejecutivo-contrato", final_report_payload, timeout=120.0)
        if isinstance(final_report, dict):
            final_report.setdefault("fecha_proceso", legacy_fecha_proceso)
            try:
                final_report.setdefault("fecha_proceso_human", _format_date_es(datetime.strptime(legacy_fecha_proceso, "%Y%m%d")))
            except ValueError:
                final_report.setdefault("fecha_proceso_human", legacy_fecha_proceso)
            final_summary = final_report.setdefault("resumen_ejecutivo", {})
            if isinstance(final_summary, dict):
                final_summary.setdefault("fecha_proceso", legacy_fecha_proceso)
                final_summary.setdefault("fecha_proceso_human", final_report.get("fecha_proceso_human", legacy_fecha_proceso))
                pre_summary = ((analysis.get("reporte_ejecutivo") or {}).get("resumen_ejecutivo") or {})
                if isinstance(pre_summary, dict):
                    final_summary.setdefault("nomina_total", pre_summary.get("nomina_total"))
                    final_report.setdefault("nomina_total", pre_summary.get("nomina_total"))
        timeline.append(
            _workflow_step(
                "reporte_final",
                "Reporte ejecutivo final",
                "ok",
                "Reporte ejecutivo final consolidado.",
                {"report": final_report},
            )
        )
    except Exception as exc:
        timeline.append(_workflow_step("reporte_final", "Reporte ejecutivo final", "failed", str(exc)))

    timeline.append(
        _workflow_step(
            "cierre_flujo",
            "Cierre del flujo",
            "ok",
            "Flujo automático completado.",
            {"lote": lote},
        )
    )
    workflow = {
        "status": "completed",
        "current_step": "cierre_flujo",
        "steps": timeline,
        "stop_reason": "",
        "executive_report_precheck": analysis.get("reporte_ejecutivo"),
        "executive_report_final": final_report,
        "output_926": output_926,
    }
    payload["analysis"]["output_926"] = output_926
    payload["analysis"]["workflow_run"] = workflow
    payload["updated_at"] = utc_now()
    save_case(payload)
    return payload


def analyze_case(case_id: str) -> Dict[str, Any]:
    payload = load_case(case_id)
    files = payload.get("files", [])
    xlsx_profile: Dict[str, Any] = {}
    docs: List[Dict[str, Any]] = []
    xlsx_entry: Optional[Dict[str, Any]] = None
    clean_output: Dict[str, Any] = {}

    for file_entry in files:
        path = Path(file_entry["stored_path"])
        suffix = path.suffix.lower()
        if suffix in {".xlsx", ".xlsm", ".xls"}:
            xlsx_entry = file_entry
            xlsx_profile = _read_xlsx(path)
            continue

        if suffix == ".pdf":
            result = _read_pdf(path)
        elif suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"}:
            result = _ocr_image(path)
        else:
            text = normalize_text(path.read_text(encoding="utf-8", errors="ignore")) if path.exists() else ""
            result = {"text": text, "used_ocr": False, "pages_processed": 1 if text else 0}

        doc_meta = _classify_document(path.name, result["text"])
        fields = _extract_fields(result["text"])
        docs.append(
            {
                "filename": path.name,
                "document_type": doc_meta["document_type"],
                "legacy_code": doc_meta["legacy_code"],
                "legacy_label": LEGACY_CODE_TO_TYPE.get(doc_meta["legacy_code"], ""),
                "code_source": doc_meta["code_source"],
                "used_ocr": result["used_ocr"],
                "pages_processed": result["pages_processed"],
                "ocr_text": result["text"],
                "text_preview": result["text"][:600],
                "fields": fields,
            }
        )

    _apply_document_classification_overrides(docs)

    if xlsx_entry:
        xlsx_bytes = Path(xlsx_entry["stored_path"]).read_bytes()
        clean_output = _generate_clean_via_legacy_nova(xlsx_entry["filename"], xlsx_bytes)
        if not bool(clean_output.get("ok")) and xlsx_bytes:
            workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
            clean_output = _generate_clean_from_workbook(workbook, xlsx_entry["filename"])
            clean_output["_source"] = "local_fallback"
        xlsx_profile = _enrich_xlsx_profile_from_clean(xlsx_profile, clean_output, docs)
        xlsx_profile = _finalize_profile_from_docs(xlsx_profile, docs)

    required_docs = _build_required_documents(xlsx_profile.get("profile", {}))
    received_types = {item["document_type"] for item in docs}
    required_evidence = _infer_required_document_satisfaction(required_docs, docs, xlsx_profile.get("profile", {}))
    missing_docs = [doc for doc in required_docs if not (required_evidence.get(doc) or {}).get("satisfied")]

    xlsx_document = only_digits(xlsx_profile.get("profile", {}).get("documento", ""))
    matched_docs = [
        item["filename"]
        for item in docs
        if xlsx_document and xlsx_document == only_digits(item["fields"].get("document_number", ""))
    ]
    mismatches = []
    if xlsx_document and not matched_docs:
        mismatches.append("Ningun documento OCR coincide con el documento principal reportado en el XLSX.")

    blockers = []
    blockers.extend(mismatches)

    validation_summary = _build_validation_summary(xlsx_profile, docs, missing_docs)
    blockers.extend(
        item["message"]
        for item in validation_summary.get("precheck", {}).get("motivos_de_rechazo", [])
        if item["message"] not in blockers
    )
    non_blocking_alerts = [
        item["message"]
        for item in validation_summary.get("alerts", [])
        if normalize_haystack(item.get("severity", "")).lower() == "alert"
    ]
    blockers.extend(
        item["message"]
        for item in validation_summary.get("alerts", [])
        if normalize_haystack(item.get("severity", "")).lower() != "alert" and item["message"] not in blockers
    )

    decision_status = "aprobable" if not blockers else "observado"
    next_step = (
        "Validar expediente final y radicar afiliacion."
        if decision_status == "aprobable"
        else "Solicitar faltantes o corregir inconsistencias antes de radicar."
    )

    checklist = {
        "required": required_docs,
        "received": sorted(received_types),
        "missing": missing_docs,
        "required_evidence": required_evidence,
        "matched_documents": matched_docs,
        "mismatches": mismatches,
        "received_summary": _summarize_received_documents(docs),
    }
    decision = {
        "flow": "afiliacion_documental",
        "recommended_status": decision_status,
        "summary": "Expediente listo para radicacion." if decision_status == "aprobable" else "Expediente con faltantes o inconsistencias.",
        "blockers": blockers,
        "alerts": non_blocking_alerts,
        "next_step": next_step,
    }
    executive_report = _build_executive_report(payload.get("label", case_id), xlsx_profile, checklist, decision, validation_summary)
    output_926 = _build_926_output(case_id, xlsx_profile, checklist, decision)

    analysis = {
        "updated_at": utc_now(),
        "xlsx_profile": xlsx_profile,
        "clean_output": {
            "source": clean_output.get("_source", ""),
            "ok": bool(clean_output),
        } if clean_output else None,
        "documents": docs,
        "checklist": checklist,
        "validacion_resumen": validation_summary,
        "decision": decision,
        "reporte_ejecutivo": executive_report,
        "draft_926": output_926.get("draft"),
        "output_926": output_926,
    }
    payload["status"] = "analyzed"
    payload["updated_at"] = utc_now()
    payload["analysis"] = analysis
    save_case(payload)
    rebuild_document_registry()
    return payload


def store_case_files(label: str, uploads: List[tuple[str, bytes]]) -> Dict[str, Any]:
    case_id = f"case-{uuid.uuid4().hex[:10]}"
    case_dir = get_case_dir(case_id)
    files_dir = case_dir / "files"
    files_dir.mkdir(parents=True, exist_ok=True)
    stored_files: List[Dict[str, Any]] = []
    seen_hashes: set[str] = set()

    def _store_one(filename: str, content: bytes) -> None:
        content_hash = hashlib.sha256(content).hexdigest()
        if content_hash in seen_hashes:
            return
        seen_hashes.add(content_hash)
        safe_name = filename or f"archivo-{uuid.uuid4().hex[:6]}"
        safe_name = safe_name.replace("\\", "/").split("/")[-1] or f"archivo-{uuid.uuid4().hex[:6]}"
        target = files_dir / safe_name
        counter = 1
        while target.exists():
            target = files_dir / f"{target.stem}-{counter}{target.suffix}"
            counter += 1
        target.write_bytes(content)
        stored_files.append(
            {
                "filename": target.name,
                "stored_path": str(target),
                "size_bytes": len(content),
                "content_type": target.suffix.lower(),
            }
        )

    def _store_processed(filename: str, content: bytes) -> None:
        lower_name = filename.lower()
        if lower_name.endswith(".pdf"):
            exploded = _explode_multipage_pdf_bytes(filename, content)
            if len(exploded) > 1:
                for page_name, page_content in exploded:
                    _store_one(page_name, page_content)
                return
        _store_one(filename, content)

    for original_name, content in uploads:
        filename = original_name or f"archivo-{uuid.uuid4().hex[:6]}"
        lower_name = filename.lower()
        if lower_name.endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(content)) as archive:
                    members = [item for item in archive.infolist() if not item.is_dir()]
                    for member in members:
                        member_name = member.filename.replace("\\", "/")
                        if member_name.startswith("__MACOSX/"):
                            continue
                        extracted = archive.read(member)
                        nested_name = member_name.split("/")[-1]
                        if not nested_name:
                            continue
                        _store_processed(nested_name, extracted)
                continue
            except zipfile.BadZipFile:
                pass
        _store_processed(filename, content)

    payload = {
        "id": case_id,
        "label": normalize_text(label) or case_id,
        "status": "uploaded",
        "created_at": utc_now(),
        "updated_at": utc_now(),
        "files": stored_files,
        "analysis": None,
    }
    return save_case(payload)


def delete_case(case_id: str) -> None:
    shutil.rmtree(get_case_dir(case_id), ignore_errors=True)
